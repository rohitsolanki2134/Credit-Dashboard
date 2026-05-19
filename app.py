"""
Credit Evaluation Dashboard — Streamlit UI
==========================================
Entry point: `streamlit run app.py`

Panels:
  1. Sidebar     — Upload / manual entry / history
  2. Snapshot    — Score, category, limit
  3. Financials  — Ratios with colour signals
  4. Payment     — Ageing buckets & trend
  5. External    — News sentiment & flags
  6. Decisions   — Rule-based + AI dual output
  7. Alerts      — Red flags
  8. Explainability — Formula walkthrough
"""

from __future__ import annotations
import io
import json
import re
import time
from dataclasses import asdict
from pathlib import Path
from typing import Optional

import load_env  # noqa: F401 — loads .env before config reads os.getenv

import numpy as np
import pandas as pd
import streamlit as st
import plotly.graph_objects as go
import plotly.express as px

# Internal modules
from config import (
    APP_ICON, APP_TITLE, PAGE_LAYOUT,
    SCORING_WEIGHTS, RISK_COLORS, RISK_EMOJI,
    DEFAULT_EMAIL_RECIPIENTS,
)
from email_sender import build_cfo_email, send_result
from database import init_db, save_evaluation, list_evaluations, get_evaluation, search_evaluations
from ingestion import ingest, ingest_payment_csv
from scoring import run_scoring, scoring_result_to_dict
from ai_engine import run_ai_analysis
from external_data import fetch_external_risk
from credit_engine import compute_credit_decision, credit_decision_to_dict
from utils import fmt_currency, fmt_pct, fmt_ratio, get_logger, score_to_band

log = get_logger(__name__)

# ---------------------------------------------------------------------------
# Page config — MUST be the first Streamlit call
# ---------------------------------------------------------------------------

st.set_page_config(
    page_title=APP_TITLE,
    page_icon=APP_ICON,
    layout=PAGE_LAYOUT,
    initial_sidebar_state="expanded",
)

# ---------------------------------------------------------------------------
# Custom CSS
# ---------------------------------------------------------------------------

st.markdown("""
<style>
/* ════════════════════════════════════════════════════════════════
   TravelPlus Credit Dashboard  ·  VeloScale structural design
   Colours: navy #0F172A · blue #2274E0 · light-blue #DBEAFE
            bg #F1F5F9 · white #FFFFFF · border #E2E8F0
   ═══════════════════════════════════════════════════════════════ */

  /* ── Reset & global ─────────────────────────────────────────── */
  .stApp { background: #F1F5F9; }
  body, .stApp { color: #0F172A; font-family: Arial, Helvetica, sans-serif; font-size: 13px; line-height: 1.4; }

  /* Main content area white background */
  .main .block-container {
      background: #F1F5F9;
      padding-top: 1.5rem !important;
  }

  /* ══════════════════════════════════════════════════════════════
     SIDEBAR — dark navy, high-contrast text throughout
     ══════════════════════════════════════════════════════════════ */
  [data-testid="stSidebar"] {
      background: #0F172A !important;
      border-right: 2px solid #2274E0 !important;
  }

  /* Every text node inside the sidebar gets readable light color */
  [data-testid="stSidebar"],
  [data-testid="stSidebar"] p,
  [data-testid="stSidebar"] span,
  [data-testid="stSidebar"] div,
  [data-testid="stSidebar"] label,
  [data-testid="stSidebar"] small,
  [data-testid="stSidebar"] strong,
  [data-testid="stSidebar"] em,
  [data-testid="stSidebar"] li,
  [data-testid="stSidebar"] .stMarkdown,
  [data-testid="stSidebar"] [data-testid="stMarkdownContainer"] { color: #CBD5E1 !important; }

  /* Headings — brightest white */
  [data-testid="stSidebar"] h1,
  [data-testid="stSidebar"] h2,
  [data-testid="stSidebar"] h3,
  [data-testid="stSidebar"] h4,
  [data-testid="stSidebar"] [data-testid="stHeading"] { color: #F8FAFC !important; }

  /* Radio button labels */
  [data-testid="stSidebar"] [data-testid="stRadio"] label,
  [data-testid="stSidebar"] [data-testid="stRadio"] p,
  [data-testid="stSidebar"] [role="radiogroup"] label { color: #E2E8F0 !important; font-weight: 500 !important; }

  /* Selected radio option — bright white */
  [data-testid="stSidebar"] [aria-checked="true"] ~ div,
  [data-testid="stSidebar"] [data-testid="stRadio"] [aria-checked="true"] + label { color: #FFFFFF !important; }

  /* Selectbox, text inputs — dark bg with light text */
  [data-testid="stSidebar"] .stSelectbox > div > div,
  [data-testid="stSidebar"] .stSelectbox [data-baseweb="select"] > div,
  [data-testid="stSidebar"] .stTextInput > div > div > input,
  [data-testid="stSidebar"] [data-baseweb="input"] {
      background: #1E293B !important;
      border-color: #334155 !important;
      color: #F1F5F9 !important;
  }
  [data-testid="stSidebar"] .stSelectbox [data-baseweb="select"] span,
  [data-testid="stSidebar"] .stSelectbox [data-baseweb="select"] div { color: #F1F5F9 !important; }

  /* File uploader */
  [data-testid="stSidebar"] [data-testid="stFileUploader"] label,
  [data-testid="stSidebar"] [data-testid="stFileUploader"] p,
  [data-testid="stSidebar"] [data-testid="stFileUploaderDropzone"] {
      color: #CBD5E1 !important;
      border-color: #334155 !important;
  }

  /* Number inputs */
  [data-testid="stSidebar"] [data-testid="stNumberInput"] input { background: #1E293B !important; border-color: #334155 !important; color: #F1F5F9 !important; }

  /* Expanders in sidebar */
  [data-testid="stSidebar"] [data-testid="stExpander"] { background: #1E293B !important; border-color: #334155 !important; }
  [data-testid="stSidebar"] [data-testid="stExpander"] summary { background: #243147 !important; color: #E2E8F0 !important; }

  /* Sidebar section divider bars */
  [data-testid="stSidebar"] hr { border-color: #2274E0 !important; opacity: 0.4; }

  /* Alert / info boxes in sidebar */
  [data-testid="stSidebar"] [data-testid="stAlert"] { background: #1E293B !important; border-left-color: #2274E0 !important; color: #CBD5E1 !important; }

  /* Sidebar buttons */
  [data-testid="stSidebar"] .stButton > button { background: #1E293B !important; color: #E2E8F0 !important; border: 1.5px solid #334155 !important; }
  [data-testid="stSidebar"] .stButton > button:hover { background: #2274E0 !important; color: #FFFFFF !important; border-color: #2274E0 !important; }

  /* Dropdown option text override */
  [data-testid="stSidebar"] [role="option"] { background: #1E293B; color: #E2E8F0; }

  /* Caption and small text */
  [data-testid="stSidebar"] .stCaption, [data-testid="stSidebar"] [data-testid="stCaptionContainer"] { color: #94A3B8 !important; }

  /* Sidebar title/logo area */
  [data-testid="stSidebarHeader"] { background: #0F172A !important; border-bottom: 1px solid #1E293B !important; }

  /* ── PANEL CARDS (VeloScale .chart-card / .flag-box style) ─── */
  .panel-card {
      background: #FFFFFF;
      border-radius: 10px;
      padding: 16px 20px;
      margin-bottom: 16px;
      box-shadow: 0 1px 6px rgba(0,0,0,.10);
      border: 1px solid #E2E8F0;
  }
  .panel-card-blue {
      background: #FFFFFF;
      border-radius: 10px;
      padding: 16px 20px;
      margin-bottom: 16px;
      box-shadow: 0 1px 6px rgba(0,0,0,.10);
      border: 1px solid #E2E8F0;
      border-top: 4px solid #2274E0;
  }
  .panel-card-red {
      background: #FFFFFF;
      border-radius: 10px;
      padding: 16px 20px;
      margin-bottom: 16px;
      box-shadow: 0 1px 6px rgba(220,38,38,.08);
      border: 1px solid #FECACA;
      border-top: 4px solid #DC2626;
  }
  .panel-card-green {
      background: #FFFFFF;
      border-radius: 10px;
      padding: 16px 20px;
      margin-bottom: 16px;
      box-shadow: 0 1px 6px rgba(22,163,74,.08);
      border: 1px solid #BBF7D0;
      border-top: 4px solid #16A34A;
  }
  .panel-card-amber {
      background: #FFFFFF;
      border-radius: 10px;
      padding: 16px 20px;
      margin-bottom: 16px;
      box-shadow: 0 1px 6px rgba(217,119,6,.08);
      border: 1px solid #FDE68A;
      border-top: 4px solid #D97706;
  }

  /* ── SECTION HEADER — VeloScale .sec style ───────────────────
     Light-blue bg pill with left accent border — very readable  */
  .section-header {
      font-size: 12px;
      font-weight: 700;
      color: #0F172A;
      background: #DBEAFE;
      padding: 7px 14px;
      border-radius: 6px;
      margin-bottom: 12px;
      border-left: 4px solid #2274E0;
      letter-spacing: 0.02em;
  }
  /* Sub-panel title — lighter version */
  .panel-title {
      font-size: 11px;
      font-weight: 700;
      color: #0F172A;
      background: #F1F5F9;
      padding: 5px 10px;
      border-radius: 5px;
      margin-bottom: 10px;
      border-left: 3px solid #94A3B8;
      text-transform: uppercase;
      letter-spacing: 0.04em;
  }

  /* ── KPI / METRIC CARDS — VeloScale .kpi style ───────────────*/
  .metric-card {
      background: #FFFFFF;
      border-radius: 10px;
      padding: 14px 16px;
      border-top: 4px solid #2274E0;
      box-shadow: 0 1px 6px rgba(0,0,0,.10);
  }
  .metric-card .kpi-label {
      font-size: 10px;
      font-weight: 700;
      text-transform: uppercase;
      letter-spacing: .5px;
      color: #64748B;
      margin-bottom: 4px;
  }
  .metric-card .kpi-val {
      font-size: 22px;
      font-weight: 700;
      color: #0F172A;
  }

  /* Native st.metric containers */
  [data-testid="metric-container"] {
      background: #FFFFFF !important;
      border-radius: 10px !important;
      padding: 14px 16px !important;
      border-top: 4px solid #2274E0 !important;
      box-shadow: 0 1px 6px rgba(0,0,0,.10) !important;
      border-left: none !important;
      border-right: none !important;
      border-bottom: 1px solid #E2E8F0 !important;
  }

  /* ── SCORE RING ──────────────────────────────────────────────── */
  .score-ring { text-align: center; font-size: 3rem; font-weight: 800; color: #0F172A; }

  /* ── FLAG / SIGNAL / CONDITION rows ─────────────────────────── */
  .red-flag {
      background: #FEE2E2;
      border-left: 4px solid #DC2626;
      border: 1px solid #FECACA;
      border-left: 4px solid #DC2626;
      padding: 8px 12px;
      border-radius: 6px;
      margin: 5px 0;
      color: #991B1B;
      font-size: 12px;
  }
  .pos-signal {
      background: #DCFCE7;
      border-left: 4px solid #16A34A;
      border: 1px solid #BBF7D0;
      border-left: 4px solid #16A34A;
      padding: 8px 12px;
      border-radius: 6px;
      margin: 5px 0;
      color: #14532D;
      font-size: 12px;
  }
  .cond-item {
      background: #DBEAFE;
      border-left: 4px solid #2274E0;
      border: 1px solid #BFDBFE;
      border-left: 4px solid #2274E0;
      padding: 8px 12px;
      border-radius: 6px;
      margin: 5px 0;
      color: #1D4ED8;
      font-size: 12px;
  }

  /* ── CHIP / BADGE — VeloScale .chip style ────────────────────── */
  .tp-badge {
      display: inline-block;
      padding: 2px 8px;
      border-radius: 20px;
      font-size: 10px;
      font-weight: 700;
      background: #DBEAFE;
      color: #1D4ED8;
      white-space: nowrap;
  }
  .tp-badge-amber { background: #FEF3C7; color: #92400E; }
  .tp-badge-green { background: #DCFCE7; color: #14532D; }
  .tp-badge-red   { background: #FEE2E2; color: #991B1B; }

  /* ── METRIC VALUES ───────────────────────────────────────────── */
  div[data-testid="stMetricValue"] { font-size: 22px !important; font-weight: 700 !important; color: #0F172A !important; }
  div[data-testid="stMetricLabel"] { font-size: 10px !important; font-weight: 700 !important; text-transform: uppercase !important; letter-spacing: .5px !important; color: #64748B !important; }
  div[data-testid="stMetricDelta"] > div { font-size: 11px !important; }

  /* ── TAB STRIP — sticky, white bar, navy active ──────────────── */
  .stTabs [data-baseweb="tab-list"] {
      background: #FFFFFF;
      border-bottom: 2px solid #DBEAFE;
      padding: 0 8px;
      gap: 2px;
  }
  .stTabs [data-baseweb="tab"] {
      height: 40px;
      padding: 0 14px;
      border-radius: 0;
      font-size: 12px;
      font-weight: 600;
      color: #1E293B;
      background: transparent;
      border: none;
      border-bottom: 3px solid transparent;
      margin-bottom: -2px;
  }
  .stTabs [aria-selected="true"] {
      color: #0F172A !important;
      font-weight: 700 !important;
      border-bottom: 3px solid #2274E0 !important;
      background: transparent !important;
  }
  .stTabs [data-testid="stTabsContent"] {
      background: #FFFFFF;
      border: 1px solid #E2E8F0;
      border-top: none;
      border-radius: 0 0 10px 10px;
      padding: 20px 16px;
  }

  /* ── EXPANDER ─────────────────────────────────────────────────── */
  [data-testid="stExpander"] {
      background: #FFFFFF !important;
      border: 1px solid #E2E8F0 !important;
      border-radius: 10px !important;
      overflow: hidden !important;
  }
  [data-testid="stExpander"] summary {
      background: #F8FAFC !important;
      font-weight: 700 !important;
      font-size: 12px !important;
      color: #0F172A !important;
      border-bottom: 1px solid #E2E8F0;
      padding: 10px 14px !important;
  }

  /* ── DIVIDER ──────────────────────────────────────────────────── */
  hr { border: none !important; border-top: 1px solid #E2E8F0 !important; margin: 12px 0 !important; }

  /* ── ALL BUTTONS — black text, never white ───────────────────── */
  .stButton > button,
  [data-testid^="stBaseButton"] { color: #0F172A !important; font-weight: 600 !important; }
  .stButton > button {
      border-radius: 6px !important;
      font-size: 12px !important;
      color: #0F172A !important;
  }
  /* Primary — amber CTA */
  .stButton > button[kind="primary"],
  [data-testid="stBaseButton-primary"] {
      background: #EAB308 !important;
      color: #0F172A !important;
      border: none !important;
      border-radius: 20px !important;
      font-weight: 700 !important;
      padding: 7px 20px !important;
  }
  .stButton > button[kind="primary"]:hover { background: #CA8A04 !important; color: #0F172A !important; }
  /* Secondary */
  .stButton > button[kind="secondary"],
  [data-testid="stBaseButton-secondary"] {
      background: #FFFFFF !important;
      color: #0F172A !important;
      border: 1.5px solid #CBD5E1 !important;
  }
  .stButton > button[kind="secondary"]:hover {
      border-color: #2274E0 !important;
      color: #2274E0 !important;
      background: #EFF6FF !important;
  }

  /* ── ALERTS ───────────────────────────────────────────────────── */
  [data-testid="stAlert"] { border-radius: 6px !important; }
  [data-testid="stAlert"][data-baseweb="notification"] { border-radius: 6px !important; }

  /* ── SCORE BAR TRACK ─────────────────────────────────────────── */
  .sub-score-row {
      background: #F8FAFC;
      border: 1px solid #E2E8F0;
      border-radius: 6px;
      padding: 8px 12px;
      margin: 4px 0;
  }

  /* ── COLUMN CONTAINERS as panel cards ────────────────────────────
     Streamlit column [data-testid="column"] is a real DOM node so
     CSS styling works perfectly here (unlike open/close st.markdown divs) */
  .block-container [data-testid="column"] {
      background: #FFFFFF;
      border-radius: 10px;
      border: 1px solid #E2E8F0;
      padding: 14px 16px !important;
      box-shadow: 0 1px 5px rgba(0,0,0,.07);
  }
  /* Metric containers already have their own card style — don't double-border */
  .block-container [data-testid="column"] [data-testid="metric-container"] {
      box-shadow: none !important;
      border-radius: 0 !important;
      border-top: none !important;
      border-left: none !important;
      border-right: none !important;
      border-bottom: 1px solid #F1F5F9 !important;
      padding: 8px 0 !important;
      background: transparent !important;
  }

  /* ── SIDEBAR FILE UPLOADER button ───────────────────────────────── */
  [data-testid="stSidebar"] [data-testid="stFileUploaderDropzone"] button,
  [data-testid="stSidebar"] [data-testid="baseButton-secondary"],
  [data-testid="stSidebar"] button[kind="secondary"],
  [data-testid="stSidebar"] [data-testid="stFileUploader"] button {
      background: #2274E0 !important;
      color: #FFFFFF !important;
      border: none !important;
      font-weight: 600 !important;
  }
  [data-testid="stSidebar"] [data-testid="stFileUploaderDropzone"] {
      background: #1E293B !important;
      border: 1.5px dashed #334155 !important;
      border-radius: 8px !important;
  }
  [data-testid="stSidebar"] [data-testid="stFileUploaderDropzone"] *,
  [data-testid="stSidebar"] [data-testid="stFileUploaderDropzone"] p,
  [data-testid="stSidebar"] [data-testid="stFileUploaderDropzone"] span {
      color: #CBD5E1 !important;
  }
  /* Sidebar: all button text must be readable */
  [data-testid="stSidebar"] button { color: #E2E8F0 !important; }
  [data-testid="stSidebar"] [data-testid="stBaseButton-secondary"] { color: #E2E8F0 !important; background: #1E293B !important; }
  [data-testid="stSidebar"] [data-testid="stBaseButton-primary"]   { color: #FFFFFF !important; background: #2274E0 !important; border: none !important; }

  /* ── SECTION SEPARATOR line ──────────────────────────────────── */
  .sec-divider { border: none; border-top: 1px solid #E2E8F0; margin: 14px 0; }
</style>
""", unsafe_allow_html=True)

# ---------------------------------------------------------------------------
# Session state initialisation
# ---------------------------------------------------------------------------

def _init_state():
    defaults = {
        "result":           None,   # last full evaluation result dict
        "financials":       None,
        "payment_records":  [],
        "company_name":     "",
        "parse_confidence": 1.0,
        "parse_warnings":   [],
        "running":          False,
        # Two-step image flow
        "awaiting_review":      False,
        "preview_financials":   None,
        "preview_confidence":   1.0,
        "preview_warnings":     [],
        "preview_scale":        10_000_000,
        "preview_company":      "",
        "preview_pay_bytes":    None,
        "preview_pay_filename": None,
        "preview_extra_fields": {},
    }
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v

_init_state()
init_db()


# ---------------------------------------------------------------------------
# Helper: gauge chart
# ---------------------------------------------------------------------------

def _gauge(score: float, title: str = "Credit Score") -> go.Figure:
    band, color, _ = score_to_band(score)
    fig = go.Figure(go.Indicator(
        mode="gauge+number+delta",
        value=score,
        delta={"reference": 50, "valueformat": ".1f"},
        title={"text": title, "font": {"size": 18}},
        gauge={
            "axis": {"range": [0, 100], "tickwidth": 1, "tickcolor": "#94A3B8"},
            "bar":  {"color": color, "thickness": 0.25},
            "steps": [
                {"range": [0,  45],  "color": "#FEE2E2"},
                {"range": [45, 70],  "color": "#FEF9C3"},
                {"range": [70, 100], "color": "#DCFCE7"},
            ],
            "threshold": {
                "line": {"color": "#0F172A", "width": 3},
                "thickness": 0.75,
                "value": score,
            },
        },
        number={"font": {"size": 40, "color": color}, "suffix": "/100"},
    ))
    fig.update_layout(
        height=280, margin=dict(l=20, r=20, t=40, b=10),
        paper_bgcolor="#FFFFFF", font_color="#0F172A",
    )
    return fig


# ---------------------------------------------------------------------------
# Helper: ratio signal → styled badge
# ---------------------------------------------------------------------------

_SIGNAL_STYLE = {
    "good":     "background:#F0FDF4;color:#16A34A;border-radius:20px;padding:2px 10px;font-size:0.78rem;font-weight:600",
    "ok":       "background:#F0F7FF;color:#2274E0;border-radius:20px;padding:2px 10px;font-size:0.78rem;font-weight:600",
    "warning":  "background:#FFFBEB;color:#D97706;border-radius:20px;padding:2px 10px;font-size:0.78rem;font-weight:600",
    "critical": "background:#FEF2F2;color:#DC2626;border-radius:20px;padding:2px 10px;font-size:0.78rem;font-weight:600",
}

def _badge(signal: str) -> str:
    style = _SIGNAL_STYLE.get(signal, _SIGNAL_STYLE["ok"])
    return f'<span style="{style}">{signal.upper()}</span>'


# ---------------------------------------------------------------------------
# Helper: generate downloadable payment history sample Excel
# ---------------------------------------------------------------------------

def _payment_sample_excel() -> bytes:
    """Return bytes of a sample payment history Excel workbook."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    # ── Sheet 1: Template ──────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Payment History"

    headers = [
        "invoice_number", "invoice_date", "due_date", "paid_date",
        "amount_inr", "currency", "description", "company_name",
    ]
    col_widths = [18, 16, 16, 16, 18, 12, 32, 28]

    header_fill = PatternFill("solid", fgColor="4F46E5")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    ws.row_dimensions[1].height = 22

    # Sample rows
    sample_rows = [
        ["INV-001", "2025-01-05", "2025-02-04", "2025-02-04",  150000, "INR", "Air travel — Q3 advance",         "Acme Corp Ltd"],
        ["INV-002", "2025-02-10", "2025-03-12", "2025-03-18",  320000, "INR", "Hotel bookings — Feb 2025",        "Acme Corp Ltd"],
        ["INV-003", "2025-03-01", "2025-03-31", "2025-04-20",  275000, "INR", "International flights — March",    "Acme Corp Ltd"],
        ["INV-004", "2025-04-08", "2025-05-08", "2025-05-07",  195000, "INR", "Domestic travel — April",          "Acme Corp Ltd"],
        ["INV-005", "2025-05-15", "2025-06-14", "2025-07-10",  410000, "INR", "Conferences & accommodation",      "Acme Corp Ltd"],
        ["INV-006", "2025-06-20", "2025-07-20", "2025-07-19",   88000, "INR", "Ground transport",                 "Acme Corp Ltd"],
        ["INV-007", "2025-07-03", "2025-08-02", "",            220000, "INR", "Pending — Aug cycle",              "Acme Corp Ltd"],
    ]

    alt_fill = PatternFill("solid", fgColor="F3F4FF")
    for r_idx, row in enumerate(sample_rows, start=2):
        fill = alt_fill if r_idx % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.fill = fill
            cell.border = border
            cell.alignment = Alignment(vertical="center")
            if c_idx in (2, 3, 4):  # date columns
                cell.number_format = "YYYY-MM-DD"
            if c_idx == 5:          # amount
                cell.number_format = "#,##0"

    ws.freeze_panes = "A2"

    # ── Sheet 2: Instructions ──────────────────────────────────────────────
    wi = wb.create_sheet("Instructions")
    wi.column_dimensions["A"].width = 22
    wi.column_dimensions["B"].width = 55
    wi.column_dimensions["C"].width = 28

    wi.cell(row=1, column=1, value="Column").font = Font(bold=True, size=12)
    wi.cell(row=1, column=2, value="Description").font = Font(bold=True, size=12)
    wi.cell(row=1, column=3, value="Example").font = Font(bold=True, size=12)

    instructions = [
        ("invoice_number", "Unique reference for the invoice (any alphanumeric)",              "INV-001"),
        ("invoice_date",   "Date the invoice was raised — YYYY-MM-DD format",                  "2025-01-05"),
        ("due_date",       "Payment due date per credit terms — YYYY-MM-DD",                   "2025-02-04"),
        ("paid_date",      "Actual date payment was received — YYYY-MM-DD (blank if unpaid)",  "2025-02-04"),
        ("amount_inr",     "Invoice amount in INR (no commas, no ₹ symbol)",                   "320000"),
        ("currency",       "Currency code — leave INR for rupee invoices",                     "INR"),
        ("description",    "Short description of the travel/service (optional)",               "Hotel bookings Feb 2025"),
        ("company_name",   "Client / counterparty name (optional — for multi-client sheets)",  "Acme Corp Ltd"),
    ]

    note_fill = PatternFill("solid", fgColor="FFFBE6")
    for r_idx, (col, desc, ex) in enumerate(instructions, start=2):
        wi.cell(row=r_idx, column=1, value=col).font = Font(bold=True, color="4F46E5")
        wi.cell(row=r_idx, column=2, value=desc)
        ex_cell = wi.cell(row=r_idx, column=3, value=ex)
        ex_cell.fill = note_fill

    wi.cell(row=12, column=1, value="NOTES").font = Font(bold=True, size=11)
    notes = [
        "• Days Late is computed automatically: paid_date − due_date",
        "• Blank paid_date = unpaid invoice (treated as outstanding/overdue)",
        "• On-time: 0 days late  |  Late-30: 1–30  |  Late-60: 31–60  |  Late-90: 61–90  |  Default: >90",
        "• Amount should be the full invoice value in INR; multi-currency is for display only",
        "• Delete the sample rows before uploading — keep only the header row + your real data",
    ]
    for i, note in enumerate(notes, start=13):
        wi.cell(row=i, column=1, value=note).font = Font(color="555555", italic=True)
        wi.merge_cells(start_row=i, start_column=1, end_row=i, end_column=3)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Sidebar — data input
# ---------------------------------------------------------------------------

def render_sidebar():
    st.sidebar.title(f"{APP_ICON} {APP_TITLE}")
    st.sidebar.markdown("---")

    mode = st.sidebar.radio(
        "Input Mode",
        ["Image / File Upload", "Listed Company (Screener.in)", "Manual Entry", "History"],
        key="sidebar_mode",
    )

    if mode == "Image / File Upload":
        _sidebar_upload()
    elif mode == "Listed Company (Screener.in)":
        _sidebar_screener()
    elif mode == "Manual Entry":
        _sidebar_manual()
    else:
        _sidebar_history()


def _sidebar_upload():
    import os as _os
    from ocr_engine import list_providers, check_provider_availability
    from config import OCR_PROVIDER as _DEFAULT_OCR

    # ── OCR provider selector ────────────────────────────────────────────────
    st.sidebar.subheader("🔍 OCR Provider")
    providers = list_providers()

    # Build display options — mark unavailable ones
    _STATUS = {True: "✅", False: "❌"}
    provider_options = {
        p["name"]: f"{_STATUS[p['available']]} {p['label']}"
        for p in providers
    }
    available_names = [p["name"] for p in providers if p["available"]]

    # Default: use config value if available, else first available
    _default = _DEFAULT_OCR if _DEFAULT_OCR in available_names else (available_names[0] if available_names else list(provider_options)[0])
    _default_idx = list(provider_options.keys()).index(_default) if _default in provider_options else 0

    selected_provider = st.sidebar.selectbox(
        "OCR / Extraction engine",
        options = list(provider_options.keys()),
        format_func = lambda k: provider_options[k],
        index   = _default_idx,
        key     = "ocr_provider_select",
        help    = "Choose how the financial image is read. Vision providers extract fields directly. Text providers do OCR first, then a lite-LLM structures the output.",
    )

    # Show availability note for selected provider
    sel_meta = next((p for p in providers if p["name"] == selected_provider), None)
    if sel_meta:
        if sel_meta["available"]:
            mode_badge = "🖼️ Vision mode" if sel_meta["mode"] == "vision" else "📝 Text + LLM-lite mode"
            st.sidebar.success(f"Active — {mode_badge}")
        else:
            st.sidebar.error(f"⚠️ {sel_meta['unavailable_reason']}")

    # ── Inline API key entry if Claude is selected but key missing ──────────
    if selected_provider == "claude" and not _os.environ.get("ANTHROPIC_API_KEY"):
        raw_key = st.sidebar.text_input(
            "Anthropic API Key",
            type        = "password",
            placeholder = "sk-ant-api03-...",
            key         = "runtime_api_key",
            help        = "Get your key at console.anthropic.com → API Keys",
        )
        if raw_key and raw_key.startswith("sk-"):
            clean_key = raw_key.strip()
            _os.environ["ANTHROPIC_API_KEY"] = clean_key
            try:
                import re as _re
                env_path = Path(__file__).parent / ".env"
                env_text = env_path.read_text(encoding="utf-8")
                new_line = f"ANTHROPIC_API_KEY={clean_key}"
                if _re.search(r"^ANTHROPIC_API_KEY=", env_text, _re.MULTILINE):
                    env_text = _re.sub(r"^ANTHROPIC_API_KEY=.*$", new_line, env_text, flags=_re.MULTILINE)
                else:
                    env_text += f"\n{new_line}\n"
                env_path.write_text(env_text, encoding="utf-8")
            except Exception:
                pass
            st.sidebar.success("✅ Key saved — ready to extract")
            st.rerun()
        elif raw_key:
            st.sidebar.warning("Key should start with `sk-ant-` — check and re-enter")

    elif selected_provider == "openai" and not _os.environ.get("OPENAI_API_KEY"):
        raw_key = st.sidebar.text_input(
            "OpenAI API Key",
            type        = "password",
            placeholder = "sk-...",
            key         = "runtime_openai_key",
            help        = "Get your key at platform.openai.com → API Keys",
        )
        if raw_key and raw_key.startswith("sk-"):
            _os.environ["OPENAI_API_KEY"] = raw_key.strip()
            st.sidebar.success("✅ OpenAI key saved")
            st.rerun()

    elif selected_provider == "tesseract" and not sel_meta["available"]:
        st.sidebar.code("pip install pytesseract pillow", language="bash")
        st.sidebar.caption("Also install the [Tesseract binary](https://github.com/UB-Mannheim/tesseract/wiki) for your OS.")

    st.sidebar.markdown("---")
    st.sidebar.subheader("Upload Financial Data")
    st.sidebar.caption(
        "Upload a **single image** of your financial statement — Claude Vision reads "
        "both the current-year and prior-year columns automatically. "
        "CSV / Excel structured exports are also accepted."
    )
    company_name = st.sidebar.text_input("Company Name *", placeholder="Acme Corp Ltd")

    fin_file = st.sidebar.file_uploader(
        "Financial Statement (both years)",
        type=["png", "jpg", "jpeg", "webp", "csv", "xlsx", "xls"],
        help=(
            "Image: photo or screenshot of any page that shows P&L / balance sheet. "
            "Annual reports typically print current year and prior year side-by-side — "
            "one upload is enough for both years. CSV/Excel: structured export."
        ),
    )

    # Scale selector — only relevant for image uploads
    _img_exts = {".png", ".jpg", ".jpeg", ".webp"}
    is_image_file = fin_file and Path(fin_file.name).suffix.lower() in _img_exts
    scale_label = st.sidebar.selectbox(
        "Values reported in",
        list(_SCALE_OPTIONS.keys()),
        index=0,   # default: Crores
        key="img_scale",
        help=(
            "Tell the parser what unit the numbers in your image use. "
            "For most Indian corporate statements: Crores. "
            "If the image shows 'Figures in ₹ Lakhs' choose Lakhs, etc."
        ),
        disabled=not is_image_file,
    )
    scale_override = _SCALE_OPTIONS[scale_label]

    pay_file = st.sidebar.file_uploader(
        "Payment History (optional)",
        type=["csv", "xlsx", "xls"],
    )

    st.sidebar.markdown("**Optional overrides (₹ Cr)**")
    travel_cost_upload = st.sidebar.number_input(
        "Annual Travel Cost (Cr)", min_value=0.0, value=0.0, step=0.5,
        help="Override if not visible in the image. Leave 0 to auto-estimate.",
    )
    travel_pct_upload = st.sidebar.slider(
        "Estimate travel as % of revenue", min_value=1, max_value=4, value=2,
        format="%d%%",
        help="Used only when no travel cost is entered or extracted.",
    )
    ar_upload = st.sidebar.number_input("Accounts Receivable override (Cr)", min_value=0.0, value=0.0, step=0.5)
    ap_upload = st.sidebar.number_input("Accounts Payable override (Cr)",    min_value=0.0, value=0.0, step=0.5)

    if st.sidebar.button("Run Credit Evaluation", type="primary", use_container_width=True):
        if not fin_file:
            st.sidebar.error("Please upload a financial statement image or file first.")
            return
        if not company_name.strip():
            st.sidebar.error("Please enter the company name.")
            return

        _run_evaluation(
            fin_bytes      = fin_file.read(),
            filename       = fin_file.name,
            company_name   = company_name.strip(),
            pay_bytes      = pay_file.read() if pay_file else None,
            use_mock_news  = False,
            scale_override = scale_override,
            ocr_provider   = st.session_state.get("ocr_provider_select"),
            extra_fields   = {
                **({"travel_cost":         travel_cost_upload * _CR} if travel_cost_upload > 0 else {}),
                "travel_budget_pct":        travel_pct_upload / 100,
                **({"accounts_receivable": ar_upload * _CR} if ar_upload > 0 else {}),
                **({"accounts_payable":    ap_upload * _CR} if ap_upload > 0 else {}),
            },
        )


_CR = 10_000_000  # 1 Crore in rupees

# Scale options for image upload (user picks unit used in their statement)
_SCALE_OPTIONS: dict = {
    "Crores (₹ Cr)":       10_000_000,
    "Lakhs (₹ L)":            100_000,
    "Millions (₹ Mn)":      1_000_000,
    "Thousands (₹ K)":          1_000,
    "Units / Raw (₹)":              1,
    "Auto-detect from image":    None,   # let parser guess from text
}

# Required fields that must be present for a valid evaluation
_REQUIRED_FIELDS = {
    "revenue":             "Revenue",
    "net_profit":          "Net Profit",
    "current_assets":      "Current Assets",
    "current_liabilities": "Current Liabilities",
    "total_debt":          "Total Debt",
    "equity":              "Net Worth / Equity",
    "operating_cash_flow": "Operating Cash Flow",
}
_RECOMMENDED_FIELDS = {
    "ebitda":               "EBITDA",
    "cash_and_equivalents": "Cash & Bank",
    "accounts_receivable":  "Receivables",
    "accounts_payable":     "Payables",
    "travel_cost":          "Travel Spend",
    "interest_expense":     "Interest Expense",
}
_PY_FIELDS = {
    "revenue_py":              "Revenue PY",
    "net_profit_py":           "Net Profit PY",
    "total_debt_py":           "Total Debt PY",
    "operating_cash_flow_py":  "Operating Cash Flow PY",
}

# For each ratio: what it measures, why critical matters, what part of the
# decision it drives, and what the analyst should watch for.
_RATIO_CRITICAL_IMPACT: dict[str, dict] = {
    "current_ratio": {
        "measures":   "Ability to cover short-term liabilities from current assets",
        "why_bad":    "Current assets are insufficient to meet near-term obligations — the company may be forced into emergency financing or default on supplier/working-capital debt.",
        "decision":   "Financial Health sub-score (liquidity component); also triggers ×0.85 financial stress haircut on credit limit.",
        "watch":      "Monitor cash burn rate and upcoming debt maturities closely.",
    },
    "quick_ratio": {
        "measures":   "Immediate liquidity — current assets minus inventory vs. current liabilities",
        "why_bad":    "Even stripping out slow-moving inventory the company cannot cover current liabilities. Acute cash pressure — relies on selling stock or new credit to stay liquid.",
        "decision":   "Financial Health sub-score (liquidity); signals inability to service short-term credit facilities we extend.",
        "watch":      "Check inventory days and whether receivables are collectible.",
    },
    "debt_to_equity": {
        "measures":   "Financial leverage — total debt relative to shareholder equity",
        "why_bad":    "Excessive borrowing relative to equity base. Creditors bear disproportionate risk; limited capacity to absorb losses before equity is wiped out.",
        "decision":   "Financial Health sub-score (leverage); if Debt/Revenue also > 1× an additional ×0.88 haircut stacks on the credit limit.",
        "watch":      "Verify whether debt is short-term (refinancing risk) or long-term. Check covenant headroom.",
    },
    "interest_coverage": {
        "measures":   "EBITDA buffer over interest payment obligations",
        "why_bad":    "Earnings barely cover interest costs. Any revenue dip or margin compression could result in an interest default — a leading indicator of distress.",
        "decision":   "Major drag on Financial Health score; AI model will prominently flag debt-servicing risk and likely recommend CONDITIONAL or REJECT.",
        "watch":      "Look at whether debt is floating-rate (rate-rise risk) and upcoming refinancing dates.",
    },
    "ebitda_margin": {
        "measures":   "Core operating profitability before interest, tax, depreciation",
        "why_bad":    "Thin or negative operating margins mean the business cannot self-fund operations, growth, or debt repayment without external cash injections.",
        "decision":   "Reduces Financial Health and Operational Stability scores; cash flow multiplier will be near-minimum (0.55×) if OCF is also negative.",
        "watch":      "Identify whether margin compression is structural (pricing pressure) or one-off (exceptional costs).",
    },
    "net_margin": {
        "measures":   "Bottom-line profitability after all costs including interest and tax",
        "why_bad":    "Net losses erode the equity base and deplete cash reserves. Sustained losses risk insolvency and may trigger auditor going-concern qualifications.",
        "decision":   "Triggers a ×0.90 net-loss haircut on credit limit; AI confidence in approval drops materially. REJECT likely if margin is deeply negative.",
        "watch":      "Distinguish between operating losses and losses driven by one-off write-downs or tax adjustments.",
    },
    "cash_flow_coverage": {
        "measures":   "Operating cash flow relative to total debt outstanding",
        "why_bad":    "Insufficient cash generation to service debt from operations alone. Company depends on asset sales, fresh equity, or debt refinancing to stay solvent.",
        "decision":   "Cash flow multiplier set to minimum 0.55×, directly cutting the credit limit. Also weighs heavily in the Financial Health sub-score.",
        "watch":      "Compare OCF to scheduled debt repayments — identify if a refinancing cliff is approaching.",
    },
    "dso": {
        "measures":   "Days Sales Outstanding — average time to collect receivables",
        "why_bad":    "Very slow collections create cash-flow gaps between billing and receipt. The company may struggle to pay its own creditors on time, increasing the risk of delayed payments to us.",
        "decision":   "DSO multiplier reduced to 0.65× for DSO > 90 days, directly reducing the credit limit. Also flags collections-quality risk in AI narrative.",
        "watch":      "Check debtor concentration — a few large slow-paying clients can distort the DSO significantly.",
    },
}


def _scale_unit(div: float) -> str:
    return {10_000_000: "Cr", 100_000: "L", 1_000_000: "Mn", 1_000: "K", 1: "₹"}.get(
        int(div) if div else 1, f"/{div:,.0f}"
    )


def _render_extraction_editor() -> None:
    """
    Editable extraction preview — shown after image parsing.

    Lets the user verify OCR results, correct any wrong values, fill in any
    blanks, then click 'Confirm & Run Evaluation'.  All changes stay in the
    data_editor widget state until confirmed.
    """
    financials  = st.session_state.get("preview_financials") or {}
    confidence  = st.session_state.get("preview_confidence", 0.0)
    warnings    = st.session_state.get("preview_warnings", [])
    scale       = st.session_state.get("preview_scale", 10_000_000) or 1
    company     = st.session_state.get("preview_company", "")

    div  = scale
    unit = _scale_unit(div)

    st.markdown("---")
    st.subheader("📋 Extraction Preview — Verify & Correct Before Evaluation")
    st.caption(
        f"OCR extracted the values below. **Click any cell to edit** incorrect or missing values, "
        f"then click **Confirm & Run Evaluation**. Values are in **{unit}** (same unit as your image)."
    )

    # ── Build editable dataframe ─────────────────────────────────────────────
    ALL_FIELDS = {**_REQUIRED_FIELDS, **_RECOMMENDED_FIELDS}
    field_keys: List[str] = []
    rows_data = []
    for key, label in ALL_FIELDS.items():
        cy_raw = financials.get(key) or 0
        py_raw = financials.get(key + "_py") or 0
        cy_cr  = round(cy_raw / div, 2) if cy_raw != 0 else None
        py_cr  = round(py_raw / div, 2) if py_raw != 0 else None
        field_keys.append(key)
        rows_data.append({
            "Required": "✅" if key in _REQUIRED_FIELDS else "⚪",
            "Field":             label,
            f"Current Year ({unit})": cy_cr,
            f"Prior Year ({unit})":   py_cr,
        })

    df_init = pd.DataFrame(rows_data)
    cy_col = f"Current Year ({unit})"
    py_col = f"Prior Year ({unit})"

    edited_df = st.data_editor(
        df_init,
        use_container_width=True,
        hide_index=True,
        disabled=["Required", "Field"],
        key="preview_editor",
        column_config={
            "Required":  st.column_config.TextColumn("", width=40),
            "Field":     st.column_config.TextColumn("Field", width="medium"),
            cy_col: st.column_config.NumberColumn(
                cy_col, format="%.2f",
                help="Current / most-recent financial year value",
            ),
            py_col: st.column_config.NumberColumn(
                py_col, format="%.2f",
                help="Prior year value (optional — enables trend analysis)",
            ),
        },
    )

    # ── Parse confidence + warnings ──────────────────────────────────────────
    conf_icon = "🟢" if confidence >= 0.70 else "🟡" if confidence >= 0.40 else "🔴"
    st.caption(f"{conf_icon} OCR confidence: **{confidence*100:.0f}%**  |  Scale: **{unit}** × {div:,.0f}")

    if warnings:
        with st.expander(f"⚠️ OCR Warnings ({len(warnings)})", expanded=confidence < 0.50):
            for w in warnings:
                st.info(w)

    # ── Validate current state of editor ────────────────────────────────────
    missing_required: List[str] = []
    for i, row in edited_df.iterrows():
        key   = field_keys[i]
        cy_v  = row[cy_col]
        is_nan = cy_v is None or (isinstance(cy_v, float) and np.isnan(cy_v))
        if key in _REQUIRED_FIELDS and (is_nan or cy_v == 0):
            missing_required.append(rows_data[i]["Field"])

    if missing_required:
        st.warning(
            f"**{len(missing_required)} required field(s) still blank** — "
            f"fill them in above before running:  "
            + ", ".join(f"**{f}**" for f in missing_required)
        )

    # ── Action buttons ───────────────────────────────────────────────────────
    col_run, col_cancel = st.columns([4, 1])
    run_clicked = col_run.button(
        "✅ Confirm & Run Evaluation",
        type="primary",
        use_container_width=True,
        disabled=bool(missing_required),
    )
    cancel_clicked = col_cancel.button("✖ Cancel", use_container_width=True)

    if cancel_clicked:
        st.session_state["awaiting_review"] = False
        st.rerun()

    if run_clicked:
        # ── Convert editor values back to raw financials dict ────────────────
        updated: dict = {}
        for i, row in edited_df.iterrows():
            key  = field_keys[i]
            cy_v = row[cy_col]
            py_v = row[py_col]
            def _valid(v):
                return v is not None and not (isinstance(v, float) and np.isnan(v)) and v != 0
            if _valid(cy_v):
                updated[key] = float(cy_v) * div
            if _valid(py_v):
                updated[key + "_py"] = float(py_v) * div

        # Preserve metadata
        for k in ("company_name", "financial_year", "prior_financial_year", "travel_budget_pct"):
            if financials.get(k):
                updated[k] = financials[k]
        if company and not updated.get("company_name"):
            updated["company_name"] = company

        # Merge extra overrides (travel cost, AR/AP from sidebar)
        for k, v in (st.session_state.get("preview_extra_fields") or {}).items():
            if v:
                updated[k] = v

        # Clear review state and run pipeline
        st.session_state["awaiting_review"] = False
        _run_pipeline(
            financials    = updated,
            company_name  = company,
            pay_bytes     = st.session_state.get("preview_pay_bytes"),
            pay_filename  = st.session_state.get("preview_pay_filename"),
        )


def _sidebar_screener():
    """Sidebar panel: fetch financials live from Screener.in for a listed company."""
    st.sidebar.subheader("🔗 Listed Company — Screener.in")
    st.sidebar.caption(
        "Paste a **Screener.in URL** or type an **NSE / BSE ticker / company name** "
        "to pull financial data automatically."
    )

    # ── Input: URL or ticker ─────────────────────────────────────────────────
    screener_input = st.sidebar.text_input(
        "Screener URL or Ticker / Company Name",
        placeholder="e.g.  RELIANCE  or  https://www.screener.in/company/RELIANCE/",
        key="screener_input_val",
    )
    consolidated = st.sidebar.checkbox("Use Consolidated Financials", value=True, key="screener_consolidated")

    # ── Search or direct fetch ───────────────────────────────────────────────
    search_clicked = st.sidebar.button("🔍 Search / Fetch", use_container_width=True)

    if search_clicked and screener_input.strip():
        raw = screener_input.strip()

        # Detect direct URL → extract symbol immediately
        url_match = re.search(r"screener\.in/company/([^/]+)", raw, re.I)
        if url_match:
            symbol = url_match.group(1).upper()
            st.session_state["screener_results"]       = [{"name": symbol, "symbol": symbol}]
            st.session_state["screener_auto_symbol"]   = symbol
        else:
            # Treat as search query
            with st.sidebar.spinner("Searching Screener.in…"):
                from screener_parser import search_company
                results = search_company(raw)
            if results:
                st.session_state["screener_results"]     = results
                st.session_state["screener_auto_symbol"] = None
            else:
                st.sidebar.error(
                    "No companies found. Try the exact NSE ticker "
                    "(e.g. `INFY`, `TCS`, `RELIANCE`) or paste the Screener.in URL directly."
                )
                st.session_state["screener_results"] = []

    # ── Company selector ─────────────────────────────────────────────────────
    results = st.session_state.get("screener_results", [])
    selected_symbol = st.session_state.get("screener_auto_symbol")

    if results and not selected_symbol:
        options = {f"{r['name']} ({r['symbol']})": r["symbol"] for r in results}
        chosen_label   = st.sidebar.selectbox("Select Company", list(options.keys()), key="screener_pick")
        selected_symbol = options[chosen_label]

    # ── Payment history (optional) ───────────────────────────────────────────
    pay_file = st.sidebar.file_uploader(
        "Payment History (optional)",
        type=["csv", "xlsx", "xls"],
        key="screener_pay_file",
    )

    # ── Fetch & Evaluate ─────────────────────────────────────────────────────
    fetch_disabled = not bool(selected_symbol)
    if st.sidebar.button(
        "📊 Fetch & Evaluate",
        type="primary",
        use_container_width=True,
        disabled=fetch_disabled,
        key="screener_fetch_btn",
    ):
        with st.sidebar.spinner(f"Fetching {selected_symbol} from Screener.in…"):
            from screener_parser import fetch_financials
            financials, fetch_warnings = fetch_financials(
                selected_symbol,
                consolidated=consolidated,
            )

        if not financials or not financials.get("revenue"):
            err = fetch_warnings[0] if fetch_warnings else "Failed to extract data."
            st.sidebar.error(
                f"{err}\n\n"
                "**Tips:**\n"
                "- Use the exact NSE ticker (e.g. `RELIANCE`, `INFY`, `HDFCBANK`)\n"
                "- Try unchecking 'Consolidated' if data is unavailable\n"
                "- Paste the full Screener.in URL for precision"
            )
            return

        company_name = financials.get("company_name", selected_symbol)

        # Store for the editable preview (same two-step flow as image uploads)
        st.session_state["awaiting_review"]      = True
        st.session_state["preview_financials"]   = financials
        st.session_state["preview_confidence"]   = 0.80
        st.session_state["preview_warnings"]     = fetch_warnings
        st.session_state["preview_scale"]        = 10_000_000   # Screener always returns Crores
        st.session_state["preview_company"]      = company_name
        st.session_state["preview_pay_bytes"]    = pay_file.read() if pay_file else None
        st.session_state["preview_pay_filename"] = pay_file.name  if pay_file else None
        st.session_state["preview_extra_fields"] = {}
        st.rerun()

    # ── Helper tip ───────────────────────────────────────────────────────────
    with st.sidebar.expander("ℹ️ How to find the right ticker"):
        st.markdown("""
        1. Go to **[screener.in](https://www.screener.in)**
        2. Search for the company name in the search bar
        3. Copy the ticker from the URL:
           `screener.in/company/**RELIANCE**/consolidated/`
        4. Paste the full URL **or** just the ticker here

        **Common tickers:** `RELIANCE`, `TCS`, `INFY`, `HDFCBANK`,
        `ICICIBANK`, `WIPRO`, `BAJFINANCE`, `ADANIPORTS`
        """)


def _sidebar_manual():
    st.sidebar.subheader("Manual Financial Entry")
    company_name = st.sidebar.text_input("Company Name *", placeholder="Acme Corp Ltd")
    fy_label     = st.sidebar.text_input("Current FY Label", value="FY25", help="e.g. FY25")

    # ---- Current year ----
    st.sidebar.markdown(f"**Income Statement — {fy_label} (₹ Cr)**")
    revenue    = st.sidebar.number_input("Annual Revenue (Cr)",          min_value=0.0,        value=50.0,  step=1.0)
    ebitda     = st.sidebar.number_input("EBITDA (Cr)",                  min_value=-100_000.0, value=8.0,   step=0.5)
    net_profit = st.sidebar.number_input("Net Profit (Cr)",              min_value=-100_000.0, value=4.0,   step=0.5)

    st.sidebar.markdown(f"**Balance Sheet — {fy_label} (₹ Cr)**")
    current_assets = st.sidebar.number_input("Current Assets (Cr)",          min_value=0.0,        value=15.0,  step=1.0)
    current_liab   = st.sidebar.number_input("Current Liabilities (Cr)",     min_value=0.0,        value=7.0,   step=0.5)
    total_debt     = st.sidebar.number_input("Total Debt / Borrowings (Cr)", min_value=0.0,        value=20.0,  step=1.0)
    equity         = st.sidebar.number_input("Total Equity / Net Worth (Cr)",min_value=-100_000.0, value=18.0,  step=1.0)
    cash           = st.sidebar.number_input("Cash & Bank Balances (Cr)",    min_value=0.0,        value=3.0,   step=0.5)
    accounts_recv  = st.sidebar.number_input("Accounts Receivable (Cr)",     min_value=0.0,        value=0.0,   step=0.5,
                                             help="Trade receivables — used to compute DSO")
    accounts_pay   = st.sidebar.number_input("Accounts Payable (Cr)",        min_value=0.0,        value=0.0,   step=0.5,
                                             help="Trade payables — used to compute DPO")

    st.sidebar.markdown(f"**Cash Flow — {fy_label} (₹ Cr)**")
    op_cash_flow   = st.sidebar.number_input("Operating Cash Flow (Cr)",     min_value=-100_000.0, value=6.0,   step=0.5)
    interest_exp   = st.sidebar.number_input("Finance Costs / Interest (Cr)",min_value=0.0,        value=0.8,   step=0.1)

    st.sidebar.markdown("**Travel Spend (₹ Cr)**")
    travel_cost = st.sidebar.number_input(
        "Annual Travel Cost (Cr)",
        min_value=0.0, value=0.0, step=0.5,
        help="From P&L/notes. Leave 0 to estimate from revenue.",
    )
    travel_pct = st.sidebar.slider(
        "Estimate travel as % of revenue",
        min_value=1, max_value=4, value=2, format="%d%%",
        help="Used only when no travel cost is entered above.",
    ) if travel_cost == 0 else 2

    # ---- Prior year (optional, same fields as current year) ----
    with st.sidebar.expander("📅 Prior Year Financials (for trend analysis)"):
        fy_py_label     = st.text_input("Prior FY Label", value="FY24")
        st.markdown(f"**Income Statement — {fy_py_label} (₹ Cr)**")
        revenue_py      = st.number_input(f"Annual Revenue (Cr)",          min_value=0.0,        value=0.0, step=1.0,   key="rev_py")
        ebitda_py       = st.number_input(f"EBITDA (Cr)",                  min_value=-100_000.0, value=0.0, step=0.5,   key="ebitda_py")
        net_prof_py     = st.number_input(f"Net Profit (Cr)",              min_value=-100_000.0, value=0.0, step=0.5,   key="np_py")
        st.markdown(f"**Balance Sheet — {fy_py_label} (₹ Cr)**")
        cur_assets_py   = st.number_input(f"Current Assets (Cr)",          min_value=0.0,        value=0.0, step=1.0,   key="ca_py")
        cur_liab_py     = st.number_input(f"Current Liabilities (Cr)",     min_value=0.0,        value=0.0, step=0.5,   key="cl_py")
        total_dbt_py    = st.number_input(f"Total Debt / Borrowings (Cr)", min_value=0.0,        value=0.0, step=1.0,   key="td_py")
        equity_py       = st.number_input(f"Total Equity / Net Worth (Cr)",min_value=-100_000.0, value=0.0, step=1.0,   key="eq_py")
        cash_py         = st.number_input(f"Cash & Bank Balances (Cr)",    min_value=0.0,        value=0.0, step=0.5,   key="cash_py")
        ar_py           = st.number_input(f"Accounts Receivable (Cr)",     min_value=0.0,        value=0.0, step=0.5,   key="ar_py")
        ap_py           = st.number_input(f"Accounts Payable (Cr)",        min_value=0.0,        value=0.0, step=0.5,   key="ap_py")
        st.markdown(f"**Cash Flow — {fy_py_label} (₹ Cr)**")
        op_cf_py        = st.number_input(f"Operating Cash Flow (Cr)",     min_value=-100_000.0, value=0.0, step=0.5,   key="ocf_py")
        int_exp_py      = st.number_input(f"Finance Costs / Interest (Cr)",min_value=0.0,        value=0.0, step=0.1,   key="ie_py")

    # ---- Payment history (optional) ----
    with st.sidebar.expander("💳 Payment History (optional)"):
        st.download_button(
            label="⬇ Download Sample Template (.xlsx)",
            data=_payment_sample_excel(),
            file_name="payment_history_template.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
        pay_file_manual = st.file_uploader(
            "Upload payment history (CSV or Excel)",
            type=["csv", "xlsx", "xls"],
            key="pay_manual",
            help="Required columns: invoice_date, due_date, paid_date, amount_inr (or amount). "
                 "Use the sample template above for the correct format.",
        )

    if st.sidebar.button("Run Credit Evaluation", type="primary", use_container_width=True):
        if not company_name.strip():
            st.sidebar.error("Company name is required.")
            return

        financials = {
            "company_name":         company_name.strip(),
            "revenue":              revenue        * _CR,
            "ebitda":               ebitda         * _CR,
            "net_profit":           net_profit     * _CR,
            "current_assets":       current_assets * _CR,
            "current_liabilities":  current_liab   * _CR,
            "total_debt":           total_debt     * _CR,
            "equity":               equity         * _CR,
            "cash_and_equivalents": cash           * _CR,
            "operating_cash_flow":  op_cash_flow   * _CR,
            "interest_expense":     interest_exp   * _CR,
            "travel_budget_pct":    travel_pct / 100,
        }
        if travel_cost > 0:
            financials["travel_cost"] = travel_cost * _CR
        if accounts_recv > 0:
            financials["accounts_receivable"] = accounts_recv * _CR
        if accounts_pay > 0:
            financials["accounts_payable"] = accounts_pay * _CR
        # Prior year (only include non-zero values)
        if revenue_py > 0:
            financials["revenue_py"]    = revenue_py    * _CR
        if net_prof_py != 0:
            financials["net_profit_py"] = net_prof_py   * _CR
        if ebitda_py != 0:
            financials["ebitda_py"]     = ebitda_py     * _CR
        if total_dbt_py > 0:
            financials["total_debt_py"] = total_dbt_py  * _CR
        if equity_py != 0:
            financials["equity_py"]     = equity_py     * _CR
        if cur_assets_py > 0:
            financials["current_assets_py"]       = cur_assets_py * _CR
        if cur_liab_py > 0:
            financials["current_liabilities_py"]  = cur_liab_py   * _CR
        if cash_py > 0:
            financials["cash_and_equivalents_py"] = cash_py       * _CR
        if ar_py > 0:
            financials["accounts_receivable_py"]  = ar_py         * _CR
        if ap_py > 0:
            financials["accounts_payable_py"]     = ap_py         * _CR
        if op_cf_py != 0:
            financials["operating_cash_flow_py"]  = op_cf_py      * _CR
        if int_exp_py > 0:
            financials["interest_expense_py"]     = int_exp_py    * _CR

        _run_evaluation(
            fin_bytes         = None,
            filename          = "",
            company_name      = company_name.strip(),
            pay_bytes         = pay_file_manual.read() if pay_file_manual else None,
            pay_filename      = pay_file_manual.name if pay_file_manual else None,
            use_mock_news     = False,
            manual_financials = financials,
        )


def _sidebar_history():
    st.sidebar.subheader("Search History")
    st.sidebar.text_input(
        "Search by company name",
        key="hist_search_query",
        placeholder="Type to filter…",
    )
    st.sidebar.selectbox(
        "Filter by risk category",
        ["All", "Low Risk", "Average", "Risky"],
        key="hist_risk_filter",
    )
    st.sidebar.caption("Results appear in the main panel →")


# ---------------------------------------------------------------------------
# Evaluation pipeline
# ---------------------------------------------------------------------------

def _run_evaluation(
    fin_bytes, filename, company_name,
    pay_bytes=None, pay_filename=None, use_mock_news=False,
    manual_financials=None,
    extra_fields: dict = None,
    scale_override: float = None,
    ocr_provider: str = None,
):
    """
    Entry point from sidebar buttons.

    For image files  → parse, then store state for the editable preview.
    For manual/csv   → run full pipeline immediately.
    """
    _img_exts = {".png", ".jpg", ".jpeg", ".webp"}
    is_image = (filename and Path(filename).suffix.lower() in _img_exts
                and not manual_financials)

    if is_image:
        # ── Image path: parse → store → show editable preview ────────────────
        st.session_state["running"] = True
        progress = st.sidebar.progress(0, "Parsing image…")
        try:
            progress.progress(30, "Running OCR…")
            financials, confidence, warnings = ingest(
                fin_bytes, filename=filename, company_name=company_name,
                scale_override=scale_override, ocr_provider=ocr_provider,
            )
            progress.progress(100, "Done — verify below ✓")
            time.sleep(0.3)
        except Exception as exc:
            st.sidebar.error(f"Image parsing failed: {exc}")
            log.error("Image parse error: %s", exc, exc_info=True)
            return
        finally:
            progress.empty()
            st.session_state["running"] = False

        # Store for the editable preview
        st.session_state["awaiting_review"]      = True
        st.session_state["preview_financials"]   = financials
        st.session_state["preview_confidence"]   = confidence
        st.session_state["preview_warnings"]     = warnings
        st.session_state["preview_scale"]        = scale_override or 10_000_000
        st.session_state["preview_company"]      = company_name
        st.session_state["preview_pay_bytes"]    = pay_bytes
        st.session_state["preview_pay_filename"] = pay_filename
        st.session_state["preview_extra_fields"] = extra_fields or {}
        return  # main() will render the editor on next rerun

    # ── Non-image path: run full pipeline immediately ─────────────────────
    st.session_state["running"] = True
    progress = st.sidebar.progress(0, "Starting…")
    try:
        progress.progress(10, "Parsing financial data…")
        if manual_financials:
            from ingestion import _passthrough
            financials, confidence, warnings = _passthrough(manual_financials)
        else:
            financials, confidence, warnings = ingest(
                fin_bytes, filename=filename, company_name=company_name,
                scale_override=scale_override, ocr_provider=ocr_provider,
            )

        if extra_fields:
            for k, v in extra_fields.items():
                if v is not None:
                    financials[k] = v

        _run_pipeline(
            financials    = financials,
            company_name  = financials.get("company_name", company_name),
            pay_bytes     = pay_bytes,
            pay_filename  = pay_filename,
            use_mock_news = use_mock_news,
            confidence    = confidence,
            warnings      = warnings,
            _progress     = progress,
        )
    except Exception as exc:
        log.error("Evaluation failed: %s", exc, exc_info=True)
        st.sidebar.error(f"Evaluation failed: {exc}")
        progress.empty()
    finally:
        st.session_state["running"] = False


def _run_pipeline(
    financials: dict,
    company_name: str,
    pay_bytes=None,
    pay_filename=None,
    use_mock_news: bool = False,
    confidence: float = 1.0,
    warnings: list = None,
    _progress=None,
):
    """
    Run scoring → AI → credit decision → persist.
    Called both from _run_evaluation (non-image) and from the editor confirm button.
    """
    warnings = warnings or []
    own_progress = _progress is None
    progress = _progress or st.sidebar.progress(0, "Starting…")
    st.session_state["running"] = True

    try:
        company_name = financials.get("company_name", company_name)
        st.session_state["financials"]       = financials
        st.session_state["parse_confidence"] = confidence
        st.session_state["parse_warnings"]   = warnings
        st.session_state["company_name"]     = company_name

        # Payment history
        progress.progress(25, "Processing payment history…")
        payment_records = []
        if pay_bytes:
            try:
                payment_records = ingest_payment_csv(pay_bytes, filename=pay_filename)
            except Exception as e:
                warnings.append(f"Payment file parse error: {e}")

        # External risk
        progress.progress(40, "Fetching external risk signals…")
        external = fetch_external_risk(company_name, use_mock=use_mock_news)

        # Scoring
        progress.progress(60, "Running scoring model…")
        scoring = run_scoring(financials, payment_records, external)

        # AI
        progress.progress(75, "Running AI analysis…")
        payment_stats = {"total_invoices": len(payment_records),
                         "on_time_pct": "N/A", "avg_days_late": "N/A"}
        ai = run_ai_analysis(
            financials,
            [asdict(r) for r in scoring.ratios],
            scoring_result_to_dict(scoring),
            external,
            payment_stats,
        )

        # Credit decision
        progress.progress(88, "Computing credit limit…")
        decision = compute_credit_decision(financials, scoring, external, ai.recommendation)

        # Persist
        progress.progress(95, "Saving evaluation…")
        scoring_dict  = scoring_result_to_dict(scoring)
        decision_dict = credit_decision_to_dict(decision)
        decision_dict.update({
            "composite_score": scoring.composite_score,
            "risk_category":   scoring.risk_category,
            "credit_limit":    decision.credit_limit,
            "payment_terms":   decision.payment_terms_days,
        })

        eval_id = save_evaluation(
            company_name     = company_name,
            financials       = financials,
            payment_data     = {"records": payment_records},
            scoring_result   = scoring_dict,
            external_result  = external,
            decision         = decision_dict,
            source_file      = "",
            parse_confidence = confidence,
            ai_narrative     = ai.risk_narrative,
            ai_confidence    = ai.confidence_score,
        )

        st.session_state["result"] = {
            "eval_id":         eval_id,
            "company_name":    company_name,
            "financials":      financials,
            "payment_records": payment_records,
            "external":        external,
            "scoring":         scoring,
            "ai":              ai,
            "decision":        decision,
            "parse_confidence": confidence,
            "parse_warnings":  warnings,
        }
        progress.progress(100, "Complete!")
        time.sleep(0.4)

    except Exception as exc:
        log.error("Pipeline failed: %s", exc, exc_info=True)
        st.sidebar.error(f"Evaluation failed: {exc}")
    finally:
        progress.empty()
        st.session_state["running"] = False

    st.rerun()


# ---------------------------------------------------------------------------
# Dashboard panels
# ---------------------------------------------------------------------------

def _render_critical_summary(r: dict):
    """
    Compact critical-points panel shown on the Snapshot tab.
    Shows: critical ratio flags | red flags | top AI risks | key conditions.
    """
    scoring  = r["scoring"]
    ai       = r["ai"]
    decision = r["decision"]

    # Collect critical ratios
    crit_ratios = [
        f"**{rr.name.replace('_',' ').title()}** — {rr.display_value}"
        for rr in scoring.ratios if rr.signal == "critical"
    ]
    # Collect warning ratios (non-critical but worth flagging)
    warn_ratios = [
        f"**{rr.name.replace('_',' ').title()}** — {rr.display_value}"
        for rr in scoring.ratios if rr.signal == "warning"
    ]
    # Structural red flags
    red_flags = list(scoring.red_flags)
    # Top positive signals
    positives = list(scoring.positive_signals)[:4]
    # AI top risks (deduplicate against red_flags)
    ai_risks  = ai.key_risks[:4]
    # Conditions
    all_conds = list(decision.conditions) + ai.suggested_conditions

    st.markdown('<div class="panel-card" style="padding:0;overflow:hidden">'
                '<div style="background:#0F172A;padding:10px 20px;border-radius:11px 11px 0 0">'
                '<span style="color:#F8FAFC;font-size:0.72rem;font-weight:800;text-transform:uppercase;letter-spacing:0.1em">'
                '📋 Credit Summary — Flags · Signals · Conditions</span></div>'
                '<div style="padding:20px 20px 4px 20px">',
                unsafe_allow_html=True)
    col_alert, col_pos, col_cond = st.columns(3, gap="large")

    with col_alert:
        st.markdown('<div class="panel-title">🚨 Critical & Risk Flags</div>', unsafe_allow_html=True)

        # Critical ratios — with impact detail
        if crit_ratios:
            st.markdown("**Critical Ratios**")
            for rr in scoring.ratios:
                if rr.signal != "critical":
                    continue
                impact = _RATIO_CRITICAL_IMPACT.get(rr.name, {})
                label  = rr.name.replace("_", " ").title()
                st.markdown(
                    f'<div style="background:#FEF2F2;border-left:4px solid #DC2626;'
                    f'padding:8px 12px;border-radius:6px;margin:4px 0">'
                    f'<span style="font-weight:700;font-size:0.9rem">🔴 {label}: {rr.display_value}</span><br>'
                    f'<span style="font-size:0.78rem;color:#6B7280">📌 <b>Measures:</b> {impact.get("measures","")}</span><br>'
                    f'<span style="font-size:0.78rem;color:#D97706">⚠️ <b>Why critical:</b> {impact.get("why_bad","")}</span><br>'
                    f'<span style="font-size:0.78rem;color:#9CA3AF">⚙️ <b>Decision impact:</b> {impact.get("decision","")}</span>'
                    f'</div>',
                    unsafe_allow_html=True,
                )

        # Warning ratios (compact, max 3)
        if warn_ratios:
            st.markdown("**Warnings**")
            shown = 0
            for rr in scoring.ratios:
                if rr.signal != "warning" or shown >= 3:
                    continue
                impact = _RATIO_CRITICAL_IMPACT.get(rr.name, {})
                label  = rr.name.replace("_", " ").title()
                st.markdown(
                    f'<div style="background:#FFFBEB;border-left:3px solid #D97706;'
                    f'padding:6px 10px;border-radius:5px;margin:3px 0">'
                    f'<span style="font-weight:600;font-size:0.87rem">🟡 {label}: {rr.display_value}</span><br>'
                    f'<span style="font-size:0.76rem;color:#6B7280">{impact.get("watch","")}</span>'
                    f'</div>',
                    unsafe_allow_html=True,
                )
                shown += 1

        if red_flags:
            st.markdown("**Risk Flags**")
            for flag in red_flags[:4]:
                st.markdown(
                    f'<div class="red-flag" style="font-size:0.85rem">⚠️ {flag}</div>',
                    unsafe_allow_html=True,
                )
        if ai_risks:
            st.markdown("**AI Key Risks**")
            for risk in ai_risks:
                st.markdown(
                    f'<div class="red-flag" style="font-size:0.85rem">🤖 {risk}</div>',
                    unsafe_allow_html=True,
                )
        if not crit_ratios and not red_flags and not ai_risks and not warn_ratios:
            st.success("No critical flags — clean risk profile ✅")

    with col_pos:
        st.markdown('<div class="panel-title">✅ Positive Signals</div>', unsafe_allow_html=True)
        if ai.positive_factors:
            st.markdown("**AI Positives**")
            for pos in ai.positive_factors[:4]:
                st.markdown(
                    f'<div class="pos-signal" style="font-size:0.85rem">✅ {pos}</div>',
                    unsafe_allow_html=True,
                )
        if positives:
            st.markdown("**Strong Ratios**")
            for pos in positives:
                st.markdown(
                    f'<div class="pos-signal" style="font-size:0.85rem">📈 {pos}</div>',
                    unsafe_allow_html=True,
                )
        if not ai.positive_factors and not positives:
            st.info("No positive signals identified.")

    with col_cond:
        st.markdown('<div class="panel-title">📋 Conditions & Terms</div>', unsafe_allow_html=True)
        # Credit decision strip
        approved_icon  = "✅ APPROVED" if decision.approved else "❌ DECLINED"
        approved_color = "#16A34A" if decision.approved else "#DC2626"
        st.markdown(
            f'<div style="font-size:1.1rem;font-weight:700;color:{approved_color}'
            f';margin-bottom:6px">{approved_icon}</div>',
            unsafe_allow_html=True,
        )
        st.markdown(f"💰 **Limit:** {fmt_currency(decision.credit_limit)}")
        st.markdown(f"📅 **Terms:** Net-{decision.payment_terms_days} days")
        st.markdown(f"🛡️ **Max Exposure:** {fmt_currency(decision.max_safe_exposure)}")
        if all_conds:
            st.markdown("**Conditions:**")
            for cond in all_conds[:5]:
                st.markdown(
                    f'<div class="cond-item" style="font-size:0.82rem">📌 {cond}</div>',
                    unsafe_allow_html=True,
                )

    st.markdown('</div></div>', unsafe_allow_html=True)


def render_snapshot(r: dict):
    scoring  = r["scoring"]
    decision = r["decision"]
    ai       = r["ai"]

    band, color, emoji = score_to_band(scoring.composite_score)
    approved_label = "APPROVED" if decision.approved else "DECLINED"
    approved_color = "#16A34A" if decision.approved else "#DC2626"

    # ── Initialise dialog flag ───────────────────────────────────────────────
    if "open_email_dlg" not in st.session_state:
        st.session_state["open_email_dlg"] = False

    # ── Top header bar + email button side-by-side ──────────────────────────
    hdr_col, btn_col = st.columns([5, 1])
    with hdr_col:
        st.markdown(
            f"<div style='background:#0F172A;border-radius:12px;padding:16px 24px;"
            f"display:flex;align-items:center;justify-content:space-between'>"
            f"<span style='color:#F8FAFC;font-size:1.35rem;font-weight:800'>"
            f"{APP_ICON} {r['company_name']} — Credit Evaluation</span>"
            f"<span style='background:{approved_color};color:#FFFFFF;font-weight:700;"
            f"font-size:0.9rem;border-radius:20px;padding:4px 16px'>{approved_label}</span>"
            f"</div>",
            unsafe_allow_html=True,
        )
    with btn_col:
        st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)
        if st.button("📧 Email Report", key="open_email_dialog",
                     use_container_width=True, type="primary"):
            st.session_state["open_email_dlg"] = True

    # ── Open dialog OUTSIDE column context so it renders at page level ───────
    if st.session_state.get("open_email_dlg"):
        _email_dialog(r)

    st.markdown('<div class="panel-card-blue">', unsafe_allow_html=True)
    col_gauge, col_metrics = st.columns([1, 2], gap="large")

    with col_gauge:
        st.plotly_chart(_gauge(scoring.composite_score), use_container_width=True)
        st.markdown(
            f"<div style='text-align:center;font-size:1.5rem;font-weight:700;"
            f"color:{color}'>{emoji} {band} Risk</div>",
            unsafe_allow_html=True,
        )

    with col_metrics:
        m1, m2, m3 = st.columns(3)
        m1.metric("Composite Score", f"{scoring.composite_score:.1f}/100")
        m2.metric("Credit Limit",    fmt_currency(decision.credit_limit))
        m3.metric("Payment Terms",   f"Net-{decision.payment_terms_days}")

        m4, m5, m6 = st.columns(3)
        m4.metric("Max Safe Exposure", fmt_currency(decision.max_safe_exposure))
        m5.metric(
            "AI Decision",
            ai.recommendation,
            delta=f"{ai.confidence_level} confidence",
        )
        m6.metric("Rule Decision", band)

        # Confidence bar
        st.markdown(f"**AI Confidence: {ai.confidence_score:.0%}**")
        st.progress(ai.confidence_score)

        # Parse confidence if from file
        if r["parse_confidence"] < 1.0:
            st.info(
                f"Data Extraction Confidence: **{r['parse_confidence']:.0%}**  "
                "— please verify extracted figures below."
            )

    st.markdown('</div>', unsafe_allow_html=True)

    # Parse warnings
    if r.get("parse_warnings"):
        with st.expander("⚠️ Data Extraction Warnings", expanded=False):
            for w in r["parse_warnings"]:
                st.warning(w)

    # ── Critical Points Summary ─────────────────────────────────────────────
    _render_critical_summary(r)


def _update_env_password(new_password: str) -> tuple:
    """
    Update SMTP_PASSWORD for the current session and persist if possible.

    Returns (success: bool, persisted: bool)
      success   — password applied to live config (always True if no exception)
      persisted — password written to .env on disk (False on read-only filesystems
                  such as Streamlit Community Cloud)
    """
    import re as _re
    import config as _cfg

    # Always apply to the live config so sending works immediately this session
    _cfg.SMTP_PASSWORD = new_password
    os.environ["SMTP_PASSWORD"] = new_password

    # Attempt to persist to .env (works locally; skipped on read-only deployments)
    env_path = os.path.join(os.path.dirname(__file__), ".env")
    if os.path.exists(env_path):
        try:
            text = open(env_path, encoding="utf-8").read()
            text = _re.sub(r"(?m)^SMTP_PASSWORD=.*$", f"SMTP_PASSWORD={new_password}", text)
            open(env_path, "w", encoding="utf-8").write(text)
            log.info("SMTP_PASSWORD updated in .env")
            return True, True
        except Exception as exc:
            log.warning("Could not write .env (read-only fs?): %s", exc)
            return True, False

    return True, False


@st.dialog("📧 Email Credit Report", width="large")
def _email_dialog(r: dict):
    import re as _re, base64 as _b64
    import config as _cfg

    # ── Initialise recipients ────────────────────────────────────────────
    if "email_recipients" not in st.session_state:
        st.session_state["email_recipients"] = list(DEFAULT_EMAIL_RECIPIENTS)

    smtp_ok = bool(_cfg.SMTP_USER and _cfg.SMTP_PASSWORD)

    # ── Status bar ───────────────────────────────────────────────────────
    if smtp_ok:
        st.markdown(
            f'<div style="background:#D1FAE5;border-left:4px solid #16A34A;'
            f'border-radius:6px;padding:7px 14px;font-size:12px;color:#065F46">'
            f'✅ Sending from <strong>{_cfg.SMTP_FROM}</strong></div>',
            unsafe_allow_html=True,
        )
    else:
        st.markdown(
            '<div style="background:#FEF3C7;border-left:4px solid #D97706;'
            'border-radius:6px;padding:7px 14px;font-size:12px;color:#92400E">'
            '⚠️ SMTP not configured — update credentials below.</div>',
            unsafe_allow_html=True,
        )

    st.markdown('<hr style="border:none;border-top:1px solid #E2E8F0;margin:10px 0">',
                unsafe_allow_html=True)

    # ── Recipients ───────────────────────────────────────────────────────
    st.markdown(
        '<div style="font-size:11px;font-weight:700;text-transform:uppercase;'
        'letter-spacing:0.05em;color:#374151;margin-bottom:8px">📬 Recipients</div>',
        unsafe_allow_html=True,
    )

    to_remove = None
    for i, addr in enumerate(st.session_state["email_recipients"]):
        c_addr, c_btn = st.columns([7, 1])
        with c_addr:
            st.markdown(
                f'<div style="background:#EFF6FF;border:1px solid #BFDBFE;border-radius:20px;'
                f'padding:5px 14px;font-size:12px;color:#1E40AF;font-weight:500;margin:2px 0">'
                f'📧 &nbsp;{addr}</div>',
                unsafe_allow_html=True,
            )
        with c_btn:
            if st.button("✕", key=f"dlg_rm_{i}", help=f"Remove {addr}"):
                to_remove = addr
    if to_remove:
        st.session_state["email_recipients"].remove(to_remove)
        st.rerun()

    # Add + Reset row
    ac1, ac2, ac3 = st.columns([5, 1, 1])
    with ac1:
        new_addr = st.text_input("Add email", placeholder="name@company.com",
                                 label_visibility="collapsed", key="dlg_new_addr")
    with ac2:
        if st.button("➕ Add", key="dlg_add", use_container_width=True):
            pat = _re.compile(r"^[^\s@]+@[^\s@]+\.[^\s@]+$")
            if pat.match(new_addr or ""):
                if new_addr not in st.session_state["email_recipients"]:
                    st.session_state["email_recipients"].append(new_addr)
                    st.rerun()
                else:
                    st.toast("Already in the list.", icon="ℹ️")
            else:
                st.toast("Enter a valid email address.", icon="⚠️")
    with ac3:
        if st.button("↺", key="dlg_reset", help="Reset to defaults",
                     use_container_width=True):
            st.session_state["email_recipients"] = list(DEFAULT_EMAIL_RECIPIENTS)
            st.rerun()

    st.markdown('<hr style="border:none;border-top:1px solid #E2E8F0;margin:10px 0">',
                unsafe_allow_html=True)

    # ── Email preview ────────────────────────────────────────────────────
    with st.expander("👁️ Preview email before sending", expanded=False):
        subject, html_body = build_cfo_email(r)
        st.caption(f"Subject: {subject}")
        encoded = _b64.b64encode(html_body.encode()).decode()
        st.markdown(
            f'<iframe src="data:text/html;base64,{encoded}" width="100%" height="420px"'
            f' style="border:1px solid #E2E8F0;border-radius:8px;display:block"></iframe>',
            unsafe_allow_html=True,
        )

    # ── SMTP password update ─────────────────────────────────────────────
    with st.expander("⚙️ Update email password", expanded=not smtp_ok):
        st.caption("Password expired or changed? Update it here — saved instantly, no restart needed.")
        p1, p2 = st.columns(2)
        with p1:
            new_pw = st.text_input("New password", type="password", key="dlg_new_pw",
                                   placeholder="New SMTP password")
        with p2:
            new_pw2 = st.text_input("Confirm", type="password", key="dlg_pw2",
                                    placeholder="Re-enter password")
        if st.button("💾 Save password", key="dlg_save_pw", use_container_width=True):
            if not new_pw:
                st.error("Password cannot be empty.")
            elif new_pw != new_pw2:
                st.error("Passwords do not match — please check and retry.")
            else:
                ok, persisted = _update_env_password(new_pw)
                if ok and persisted:
                    st.toast("✅ Password saved to .env and active.", icon="✅")
                elif ok and not persisted:
                    st.toast("✅ Password active for this session.", icon="✅")
                    st.info(
                        "**Running on Streamlit Cloud?** The password change is active "
                        "for this session only. To make it permanent, go to your app's "
                        "**Settings → Secrets** and update `SMTP_PASSWORD` there.",
                        icon="ℹ️",
                    )

    st.markdown('<hr style="border:none;border-top:1px solid #E2E8F0;margin:10px 0">',
                unsafe_allow_html=True)

    # ── Send button ──────────────────────────────────────────────────────
    n_rec = len(st.session_state["email_recipients"])
    btn_label = (
        f"📤  Send to {n_rec} recipient{'s' if n_rec != 1 else ''}"
        if n_rec > 0 else "📤  Send Report"
    )

    if st.button(btn_label, key="dlg_send", type="primary",
                 use_container_width=True,
                 disabled=not smtp_ok or n_rec == 0):
        with st.spinner(f"Sending to {n_rec} recipient(s)…"):
            ok, msg = send_result(r, st.session_state["email_recipients"])

        if ok:
            # ── SUCCESS — prominent banner + toast + auto-close ──────────
            st.markdown(
                '<div style="background:#D1FAE5;border:2px solid #16A34A;border-radius:8px;'
                'padding:14px 18px;text-align:center;margin:8px 0">'
                '<div style="font-size:1.4rem">✅</div>'
                '<div style="font-size:14px;font-weight:700;color:#065F46;margin-top:4px">'
                'Email sent successfully!</div>'
                f'<div style="font-size:12px;color:#047857;margin-top:2px">{msg}</div>'
                '</div>',
                unsafe_allow_html=True,
            )
            st.toast("✅ Email sent successfully!", icon="✅")
            # Close dialog after a short pause
            import time as _time
            _time.sleep(1.5)
            st.session_state["open_email_dlg"] = False
            st.rerun()
        else:
            # ── FAILURE — clear red banner stays visible ──────────────────
            st.markdown(
                '<div style="background:#FEE2E2;border:2px solid #DC2626;border-radius:8px;'
                'padding:14px 18px;text-align:center;margin:8px 0">'
                '<div style="font-size:1.4rem">❌</div>'
                '<div style="font-size:14px;font-weight:700;color:#991B1B;margin-top:4px">'
                'Send failed</div>'
                f'<div style="font-size:12px;color:#B91C1C;margin-top:2px">{msg}</div>'
                '</div>',
                unsafe_allow_html=True,
            )


def _ratio_interpretation(name: str, signal: str, value) -> str:
    """
    Return a one-line plain-English interpretation of a ratio result,
    tailored to whether it is good / ok / warning / critical.
    """
    interp: dict[str, dict[str, str]] = {
        "current_ratio": {
            "good":     "Strong — can comfortably pay short-term bills from existing assets.",
            "ok":       "Adequate — can cover short-term liabilities, but little cushion.",
            "warning":  "Tight — may struggle to meet near-term obligations without extra cash.",
            "critical": "Danger — current assets cannot cover short-term debt. Default risk is high.",
        },
        "quick_ratio": {
            "good":     "Strong — can cover all short-term liabilities even without selling inventory.",
            "ok":       "Acceptable — liquid assets roughly cover current liabilities.",
            "warning":  "Low — relies heavily on collecting receivables quickly to stay liquid.",
            "critical": "Very low — cannot pay short-term debts from liquid assets alone. Acute cash risk.",
        },
        "debt_to_equity": {
            "good":     "Conservative leverage — equity comfortably exceeds debt. Low financial risk.",
            "ok":       "Moderate leverage — debt is manageable relative to the equity base.",
            "warning":  "Elevated — high reliance on borrowed money; vulnerable to rate rises or revenue dips.",
            "critical": "Excessive debt — creditors are exposed far beyond the equity buffer. Distress risk.",
        },
        "interest_coverage": {
            "good":     "Excellent — EBITDA covers interest costs many times over. Very safe.",
            "ok":       "Adequate — earnings cover interest with some headroom.",
            "warning":  "Thin — a modest revenue fall could make interest payments difficult.",
            "critical": "Insufficient — earnings barely cover (or don't cover) interest obligations. Default risk.",
        },
        "ebitda_margin": {
            "good":     "Healthy operating profitability — the core business generates strong cash.",
            "ok":       "Acceptable margins — the business is profitable at the operating level.",
            "warning":  "Thin margins — limited buffer to absorb cost increases or revenue softness.",
            "critical": "Near-zero or negative — the business is not generating profit from operations.",
        },
        "net_margin": {
            "good":     "Solid bottom-line profitability after all costs, interest and tax.",
            "ok":       "Modest but positive profit — the company is making money.",
            "warning":  "Very thin profit — one bad quarter could push into losses.",
            "critical": "Net losses — the company is destroying value; equity is being eroded.",
        },
        "cash_flow_coverage": {
            "good":     "Strong — operating cash flow can service total debt comfortably.",
            "ok":       "Adequate — cash generation covers a reasonable portion of debt obligations.",
            "warning":  "Weak — cash generated from operations is insufficient to service debt fully.",
            "critical": "Critical — operating cash cannot cover debt. Relies on refinancing or asset sales.",
        },
        "dso": {
            "good":     "Fast collections — receivables convert to cash quickly. Good liquidity signal.",
            "ok":       "Normal — receivables are collected within typical industry timelines.",
            "warning":  "Slow — clients are taking longer to pay; watch for cash flow gaps.",
            "critical": "Very slow collections — high risk of payment delays flowing through to us.",
        },
    }
    return interp.get(name, {}).get(signal, "—")


def _render_ratio_card(rr) -> None:
    """Render a single ratio as a compact card with value, score bar, and plain-English meaning."""
    signal = rr.signal
    colors = {
        "good":     ("#16A34A", "#F0FDF4"),
        "ok":       ("#2274E0", "#F0F7FF"),
        "warning":  ("#D97706", "#FFFBEB"),
        "critical": ("#DC2626", "#FEF2F2"),
    }
    accent, bg = colors.get(signal, ("#6B7280", "#F9FAFB"))
    label       = rr.name.replace("_", " ").title()
    score       = int(rr.normalised_score)
    meaning     = _ratio_interpretation(rr.name, signal, rr.value)
    signal_text = signal.upper()

    st.markdown(
        f'<div style="background:{bg};border-left:3px solid {accent};'
        f'padding:10px 16px;border-radius:10px;margin:5px 0;'
        f'box-shadow:0 1px 3px rgba(0,0,0,0.05)">'
        # Row 1: name + value + signal badge
        f'<div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:5px">'
        f'  <span style="font-weight:700;font-size:0.88rem;color:#111827">{label}</span>'
        f'  <span style="font-size:0.9rem;font-weight:700;color:{accent}">{rr.display_value}</span>'
        f'  <span style="background:{accent}18;color:{accent};'
        f'    border-radius:20px;padding:2px 10px;font-size:0.72rem;font-weight:700">{signal_text}</span>'
        f'</div>'
        # Row 2: score progress bar
        f'<div style="background:#F1F5F9;border:1px solid #E2E8F0;border-radius:4px;height:6px;margin-bottom:6px">'
        f'  <div style="background:{accent};width:{score}%;height:4px;border-radius:3px"></div>'
        f'</div>'
        # Row 3: plain-English meaning
        f'<div style="font-size:0.79rem;color:#6B7280;line-height:1.4">{meaning}</div>'
        f'</div>',
        unsafe_allow_html=True,
    )


def render_financial_health(r: dict):
    scoring    = r["scoring"]
    financials = r["financials"]

    st.markdown('<div class="section-header">📈 Financial Health</div>', unsafe_allow_html=True)

    # ---- YoY comparison (show when prior year available) ----
    has_py = bool(financials.get("revenue_py"))
    py_fields = [
        ("Annual Revenue",        "revenue"),
        ("EBITDA",                "ebitda"),
        ("Net Profit",            "net_profit"),
        ("Total Debt",            "total_debt"),
        ("Equity",                "equity"),
        ("Operating Cash Flow",   "operating_cash_flow"),
    ]
    if has_py:
        st.markdown('<div class="section-header">📊 Year-on-Year Comparison</div>', unsafe_allow_html=True)
        trend_rows = []
        for label, key in py_fields:
            cy = financials.get(key)
            py = financials.get(f"{key}_py")
            if cy is None and py is None:
                continue
            cy_s = fmt_currency(cy) if cy is not None else "—"
            py_s = fmt_currency(py) if py is not None else "—"
            if cy is not None and py is not None and py != 0:
                chg = (cy - py) / abs(py) * 100
                arrow = "▲" if chg >= 0 else "▼"
                chg_s = f"{arrow} {abs(chg):.1f}%"
                color = "green" if (chg >= 0 and key not in ("total_debt",)) or \
                                   (chg < 0 and key == "total_debt") else "red"
            else:
                chg_s, color = "—", "gray"
            trend_rows.append({"Metric": label, "Current Year": cy_s, "Prior Year": py_s,
                                "Change": chg_s, "_color": color})

        df_trend = pd.DataFrame(trend_rows).drop(columns=["_color"])

        def _color_change(val):
            if "▲" in str(val):
                return "color:#16A34A;font-weight:bold"
            if "▼" in str(val):
                return "color:#DC2626;font-weight:bold"
            return ""

        st.dataframe(
            df_trend.style.map(_color_change, subset=["Change"]),
            use_container_width=True, hide_index=True,
        )
        st.markdown('<hr class="sec-divider">', unsafe_allow_html=True)

    # ---- DSO / DPO metrics ----
    revenue = financials.get("revenue") or 0
    ar = financials.get("accounts_receivable") or 0
    ap = financials.get("accounts_payable") or 0
    dso = (ar / revenue * 365) if revenue > 0 and ar > 0 else None
    dpo = (ap / revenue * 365) if revenue > 0 and ap > 0 else None
    travel_c = financials.get("travel_cost")
    travel_est = revenue * (financials.get("travel_budget_pct") or 0.025)

    if dso is not None or dpo is not None or ar > 0:
        st.markdown('<div class="section-header">💰 Receivables &amp; Payables</div>', unsafe_allow_html=True)
        d1, d2, d3, d4 = st.columns(4)
        d1.metric("Accounts Receivable", fmt_currency(ar) if ar > 0 else "N/A")
        d2.metric("DSO",  f"{dso:.0f} days" if dso else "N/A",
                  delta=f"{'⚠️ High' if dso and dso > 60 else '✅ OK'}" if dso else None,
                  delta_color="off")
        d3.metric("Accounts Payable", fmt_currency(ap) if ap > 0 else "N/A")
        d4.metric("DPO",  f"{dpo:.0f} days" if dpo else "N/A")
        st.markdown('<hr class="sec-divider">', unsafe_allow_html=True)

    # ---- Travel budget ----
    st.markdown('<div class="section-header">✈️ Travel Spend Profile</div>', unsafe_allow_html=True)
    tb1, tb2, tb3 = st.columns(3)
    tb1.metric("Reported Travel Cost",
               fmt_currency(travel_c) if travel_c else "Not provided")
    tb2.metric("Estimated Travel (used)",
               fmt_currency(travel_c if travel_c else travel_est))
    tb3.metric("Travel as % of Revenue",
               f"{(travel_c or travel_est) / revenue * 100:.1f}%" if revenue > 0 else "N/A")
    st.markdown('<hr class="sec-divider">', unsafe_allow_html=True)

    # ---- Key financials table ----
    fin_data = {
        "Metric": [
            "Annual Revenue", "EBITDA", "Net Profit",
            "Total Debt", "Equity", "Operating Cash Flow",
            "Current Assets", "Current Liabilities", "Cash",
        ],
        "Value": [
            fmt_currency(financials.get("revenue")),
            fmt_currency(financials.get("ebitda")),
            fmt_currency(financials.get("net_profit")),
            fmt_currency(financials.get("total_debt")),
            fmt_currency(financials.get("equity")),
            fmt_currency(financials.get("operating_cash_flow")),
            fmt_currency(financials.get("current_assets")),
            fmt_currency(financials.get("current_liabilities")),
            fmt_currency(financials.get("cash_and_equivalents")),
        ],
    }
    col_table, col_ratios = st.columns(2, gap="large")

    with col_table:
        st.markdown('<div class="panel-title">📋 Key Financials</div>', unsafe_allow_html=True)
        st.dataframe(pd.DataFrame(fin_data), use_container_width=True, hide_index=True)

    with col_ratios:
        st.markdown('<div class="panel-title">📐 Financial Ratios</div>', unsafe_allow_html=True)
        for rr in scoring.ratios:
            _render_ratio_card(rr)

    # ── Critical / Warning ratio impact cards ──────────────────────────────
    critical_or_warn = [rr for rr in scoring.ratios if rr.signal in ("critical", "warning")]
    if critical_or_warn:
        st.markdown('<hr class="sec-divider">', unsafe_allow_html=True)
        st.markdown('<div class="section-header" style="background:#FFFBEB;border-left-color:#D97706;color:#92400E">⚠️ Ratio Impact Analysis</div>', unsafe_allow_html=True)
        st.caption(
            "Each flagged ratio below shows what it measures, why the current level is a concern, "
            "and exactly how it feeds into the credit decision."
        )
        for rr in critical_or_warn:
            impact = _RATIO_CRITICAL_IMPACT.get(rr.name, {})
            if not impact:
                continue
            label      = rr.name.replace("_", " ").title()
            is_crit    = rr.signal == "critical"
            border_col = "#DC2626" if is_crit else "#F39C12"
            bg_col     = "#FEF2F2" if is_crit else "#FFFBEB"
            icon       = "🔴" if is_crit else "🟡"
            with st.expander(
                f"{icon} {label} — {rr.display_value}  "
                f"({'Critical' if is_crit else 'Warning'}, score {rr.normalised_score:.0f}/100)",
                expanded=is_crit,
            ):
                c1, c2 = st.columns(2, gap="large")
                with c1:
                    st.markdown(
                        f'<div style="background:{bg_col};border-left:4px solid {border_col};'
                        f'padding:10px 14px;border-radius:7px;margin-bottom:8px">'
                        f'<b style="font-size:0.88rem">📌 What it measures</b><br>'
                        f'<span style="font-size:0.84rem">{impact.get("measures","—")}</span>'
                        f'</div>',
                        unsafe_allow_html=True,
                    )
                    st.markdown(
                        f'<div style="background:{bg_col};border-left:4px solid {border_col};'
                        f'padding:10px 14px;border-radius:7px">'
                        f'<b style="font-size:0.88rem">⚠️ Why this level is a concern</b><br>'
                        f'<span style="font-size:0.84rem">{impact.get("why_bad","—")}</span>'
                        f'</div>',
                        unsafe_allow_html=True,
                    )
                with c2:
                    st.markdown(
                        f'<div style="background:#F0F7FF;border-left:4px solid #2274E0;'
                        f'padding:10px 14px;border-radius:7px;margin-bottom:8px">'
                        f'<b style="font-size:0.88rem">⚙️ Decision impact</b><br>'
                        f'<span style="font-size:0.84rem">{impact.get("decision","—")}</span>'
                        f'</div>',
                        unsafe_allow_html=True,
                    )
                    st.markdown(
                        f'<div style="background:#F0F7FF;border-left:4px solid #2274E0;'
                        f'padding:10px 14px;border-radius:7px">'
                        f'<b style="font-size:0.88rem">🔍 What to watch</b><br>'
                        f'<span style="font-size:0.84rem">{impact.get("watch","—")}</span>'
                        f'</div>',
                        unsafe_allow_html=True,
                    )
    # Radar chart of sub-scores
    st.markdown('<hr class="sec-divider">', unsafe_allow_html=True)
    st.markdown('<div class="section-header">🕸️ Sub-Score Radar &amp; Ratio Scores</div>', unsafe_allow_html=True)
    sub_names   = [ss.name for ss in scoring.sub_scores]
    sub_vals    = [ss.score for ss in scoring.sub_scores]
    sub_names_c = sub_names + [sub_names[0]]
    sub_vals_c  = sub_vals  + [sub_vals[0]]

    fig_radar = go.Figure()
    fig_radar.add_trace(go.Scatterpolar(
        r=sub_vals_c, theta=sub_names_c,
        fill="toself", fillcolor="rgba(124,58,237,0.25)",
        line=dict(color="#2274E0", width=2),
        name="Sub-scores",
    ))
    fig_radar.update_layout(
        polar=dict(radialaxis=dict(visible=True, range=[0, 100])),
        showlegend=False, height=320,
        paper_bgcolor="#FFFFFF", font_color="#0F172A",
        margin=dict(l=20, r=20, t=30, b=10),
    )
    st.plotly_chart(fig_radar, use_container_width=True)
    fig_bars = px.bar(
        x=[rr.normalised_score for rr in scoring.ratios],
        y=[rr.name.replace("_", " ").title() for rr in scoring.ratios],
        orientation="h",
        color=[rr.normalised_score for rr in scoring.ratios],
        color_continuous_scale=["#DC2626", "#F39C12", "#2ECC71"],
        range_color=[0, 100],
        labels={"x": "Score (0–100)", "y": ""},
        title="Ratio Scores (0–100)",
    )
    fig_bars.update_layout(
        height=320, paper_bgcolor="#FFFFFF", font_color="#0F172A", coloraxis_showscale=False,
        margin=dict(l=10, r=10, t=40, b=10),
    )
    st.plotly_chart(fig_bars, use_container_width=True)


def render_payment_behavior(r: dict):
    records = r.get("payment_records", [])
    scoring = r["scoring"]

    st.markdown('<div class="section-header">⏱ Payment Behavior</div>', unsafe_allow_html=True)

    # Get payment sub-score
    pb_sub = next((ss for ss in scoring.sub_scores if "Payment" in ss.name), None)
    pb_score = pb_sub.score if pb_sub else 50.0

    col1, col2 = st.columns(2, gap="large")

    if not records:
        st.info("No payment history uploaded. Score defaults to neutral (50).")
        col1.metric("Payment Score", f"{pb_score:.0f}/100")
        return

    # Compute ageing buckets
    bucket_counts = {"On Time": 0, "1-30 Days": 0, "31-60 Days": 0, "61-90 Days": 0, "90+ Days": 0}
    for rec in records:
        days = rec.get("days_late", 0) or 0
        if days == 0:
            bucket_counts["On Time"] += 1
        elif days <= 30:
            bucket_counts["1-30 Days"] += 1
        elif days <= 60:
            bucket_counts["31-60 Days"] += 1
        elif days <= 90:
            bucket_counts["61-90 Days"] += 1
        else:
            bucket_counts["90+ Days"] += 1

    total = len(records)

    with col1:
        st.markdown('<div class="panel-title">📊 Payment Metrics</div>', unsafe_allow_html=True)
        st.metric("Payment Behavior Score", f"{pb_score:.0f}/100")
        st.metric("Total Invoices Analysed", total)
        pct_ontime = bucket_counts["On Time"] / total * 100
        st.metric("On-Time Payment Rate", f"{pct_ontime:.1f}%")

        avg_delay = np.mean([r.get("days_late", 0) or 0 for r in records])
        st.metric("Avg Days Late", f"{avg_delay:.0f}")

    with col2:
        st.markdown('<div class="panel-title">🕰️ Invoice Ageing Breakdown</div>', unsafe_allow_html=True)
        # Donut chart — ageing buckets
        bucket_colors = ["#16A34A", "#2274E0", "#D97706", "#F97316", "#DC2626"]
        fig_donut = go.Figure(go.Pie(
            labels=list(bucket_counts.keys()),
            values=list(bucket_counts.values()),
            hole=0.55,
            marker_colors=bucket_colors,
            textinfo="percent+label",
        ))
        fig_donut.update_layout(
            title="Invoice Ageing Buckets",
            height=300, paper_bgcolor="#FFFFFF", font_color="#0F172A", showlegend=False,
            margin=dict(l=10, r=10, t=40, b=10),
        )
        st.plotly_chart(fig_donut, use_container_width=True)

    # Table of recent payments
    if records:
        st.markdown('<hr class="sec-divider">', unsafe_allow_html=True)
        st.markdown('<div class="panel-title">📋 Recent Payment Records</div>', unsafe_allow_html=True)
        with st.expander("View Payment Records", expanded=False):
            df_pay = pd.DataFrame(records).head(20)
            st.dataframe(df_pay, use_container_width=True, hide_index=True)


def render_external_risk(r: dict):
    ext = r["external"]

    st.markdown('<div class="section-header">🌐 External Risk Signals</div>', unsafe_allow_html=True)

    sentiment = ext.get("sentiment", "Neutral")
    sent_color = {"Positive": "#16A34A", "Neutral": "#D97706", "Negative": "#DC2626"}.get(sentiment, "#888")

    # ── Top metrics strip ────────────────────────────────────────────────────
    st.markdown('<div class="section-header">🌐 Sentiment Overview</div>', unsafe_allow_html=True)
    col1, col2, col3 = st.columns(3)
    sent_icon = {"Positive": "🟢", "Neutral": "🟡", "Negative": "🔴"}.get(sentiment, "🟡")
    col1.markdown(
        f"<div style='text-align:center'>"
        f"<div style='font-size:2rem'>{sent_icon}</div>"
        f"<div style='font-size:1.4rem;font-weight:800;color:{sent_color}'>{sentiment}</div>"
        f"<div style='color:#6B7280;font-size:0.82rem'>Overall Sentiment</div>"
        f"</div>",
        unsafe_allow_html=True,
    )
    col2.metric("Risk Flags Found", len(ext.get("risk_flags", [])))
    col3.metric("Data Source", ext.get("source", "N/A").upper())
    st.markdown('<hr class="sec-divider">', unsafe_allow_html=True)

    # ── Risk flags and positive signals ─────────────────────────────────────
    col_risk, col_pos_ext = st.columns(2, gap="large")

    with col_risk:
        st.markdown('<div class="panel-title">🚩 Risk Flags</div>', unsafe_allow_html=True)
        flags_detail = ext.get("risk_flags_detail", [])
        risk_flags   = ext.get("risk_flags", [])

        if flags_detail:
            for fd in flags_detail:
                label    = fd.get("label", "Unknown")
                src_arts = fd.get("articles", [])
                # Flag header card
                st.markdown(
                    f'<div class="red-flag" style="margin:6px 0 2px 0;font-size:0.88rem">'
                    f'🚩 &nbsp;<strong>{label}</strong>'
                    f'<span style="color:#9CA3AF;font-size:0.78rem;margin-left:8px">'
                    f'— {len(src_arts)} source article{"s" if len(src_arts)!=1 else ""}</span>'
                    f'</div>',
                    unsafe_allow_html=True,
                )
                # Supporting article links
                if src_arts:
                    with st.expander("View source articles", expanded=False):
                        for art in src_arts:
                            url    = art.get("url", "")
                            title  = art.get("title") or "Article"
                            source = art.get("source", "")
                            date   = art.get("date", "")[:10]
                            link   = f"[{title}]({url})" if url else f"**{title}**"
                            st.markdown(
                                f"📰 {link}  \n"
                                f"<span style='color:#6B7280;font-size:0.80rem'>"
                                f"{source}{' &nbsp;|&nbsp; ' + date if date else ''}"
                                f"</span>",
                                unsafe_allow_html=True,
                            )
        elif risk_flags:
            # Fallback: old-style string flags (e.g. loaded from DB before this change)
            for flag in risk_flags:
                st.markdown(
                    f'<div class="red-flag" style="margin:4px 0;font-size:0.88rem">'
                    f'🚩 &nbsp;{flag}</div>',
                    unsafe_allow_html=True,
                )
        else:
            st.markdown(
                '<div class="pos-signal" style="font-size:0.88rem">'
                '✅ &nbsp;No risk flags identified</div>',
                unsafe_allow_html=True,
            )

    with col_pos_ext:
        st.markdown('<div class="panel-title">✅ Positive Signals</div>', unsafe_allow_html=True)
        pos_signals = ext.get("positive_signals", [])
        if pos_signals:
            for sig in pos_signals:
                st.markdown(
                    f'<div class="pos-signal" style="margin:4px 0;font-size:0.88rem">'
                    f'✅ &nbsp;{sig}</div>',
                    unsafe_allow_html=True,
                )
        else:
            st.markdown(
                '<div style="color:#6B7280;font-size:0.87rem;padding:5px 0">'
                'No positive signals found in recent news.</div>',
                unsafe_allow_html=True,
            )

    # ── News summary — already newline-separated bullet lines ───────────────
    # _build_summary() produces lines like:
    #   "Overall sentiment: **X**. Recent headlines:"
    #   "• Title — Source (date)"
    #   "\nRisk keywords detected: ..."
    summary_text = (ext.get("summary") or "").strip()
    if summary_text and summary_text not in ("No data available", "No recent news found for this company."):
        st.markdown('<hr class="sec-divider">', unsafe_allow_html=True)
        st.markdown('<div class="panel-title">📰 Recent News Headlines</div>', unsafe_allow_html=True)
        lines = [ln.strip() for ln in summary_text.split("\n") if ln.strip()]
        for line in lines:
            if line.startswith("•"):
                # Headline bullet
                text = line.lstrip("•").strip()
                st.markdown(
                    f'<div style="background:#F0F7FF;border-left:3px solid #2274E0;'
                    f'padding:8px 14px;border-radius:5px;margin:5px 0;font-size:0.87rem">'
                    f'📰 &nbsp;{text}</div>',
                    unsafe_allow_html=True,
                )
            elif line.lower().startswith("risk keyword"):
                # Risk keywords summary line
                st.markdown(
                    f'<div style="background:#FEF2F2;border-left:3px solid #DC2626;'
                    f'padding:7px 14px;border-radius:5px;margin:5px 0;font-size:0.84rem">'
                    f'⚠️ &nbsp;{line}</div>',
                    unsafe_allow_html=True,
                )
            else:
                # Intro / context line (e.g. "Overall sentiment: Neutral. Recent headlines:")
                st.markdown(
                    f'<div style="color:#9CA3AF;font-size:0.84rem;padding:3px 2px">'
                    f'{line}</div>',
                    unsafe_allow_html=True,
                )

    # ── Full article list (collapsed) ────────────────────────────────────────
    articles = ext.get("articles", [])
    if articles:
        st.markdown('<hr class="sec-divider">', unsafe_allow_html=True)
        with st.expander(f"📋 Full Article List ({len(articles)} articles)", expanded=False):
            for art in articles:
                sentiment_art = _score_article_sentiment(art)
                dot   = {"positive": "🟢", "negative": "🔴", "neutral": "🟡"}.get(sentiment_art, "🟡")
                url   = art.get("url", "")
                title = art.get("title", "No title")
                title_part = f"[{title}]({url})" if url else f"**{title}**"
                st.markdown(
                    f"{dot} {title_part}  \n"
                    f"<span style='color:#6B7280;font-size:0.82rem'>"
                    f"📰 {art.get('source','?')} &nbsp;|&nbsp; "
                    f"📅 {art.get('date','')[:10]}</span>",
                    unsafe_allow_html=True,
                )
                st.markdown(
                    '<hr style="border:0;border-top:1px solid #E5E7EB;margin:6px 0">',
                    unsafe_allow_html=True,
                )


def _score_article_sentiment(art: dict) -> str:
    from utils import get_logger
    from config import NEGATIVE_KEYWORDS, POSITIVE_KEYWORDS
    text = f"{art.get('title','')} {art.get('description','')}".lower()
    neg = sum(1 for kw in NEGATIVE_KEYWORDS if kw in text)
    pos = sum(1 for kw in POSITIVE_KEYWORDS if kw in text)
    if neg > pos:
        return "negative"
    if pos > neg:
        return "positive"
    return "neutral"


def render_decisions(r: dict):
    scoring  = r["scoring"]
    ai       = r["ai"]
    decision = r["decision"]

    st.markdown('<div class="section-header">🧠 Dual Decision Output</div>', unsafe_allow_html=True)

    col_rule, col_ai = st.columns(2, gap="large")

    with col_rule:
        st.markdown('<div class="section-header">🏛️ Rule-Based Decision</div>', unsafe_allow_html=True)
        band, color, emoji = score_to_band(scoring.composite_score)
        st.markdown(
            f"<div style='font-size:2rem;font-weight:800;color:{color}'>"
            f"{emoji} {band}</div>",
            unsafe_allow_html=True,
        )
        st.metric("Composite Score", f"{scoring.composite_score:.1f}/100")

        st.markdown("**Score Breakdown**")
        for label, val in scoring.score_breakdown.items():
            if label != "Composite Score":
                st.markdown(f"- {label}: **{val}**")

        st.markdown("**Sub-Score Details**")
        for ss in scoring.sub_scores:
            pct = int(ss.score)
            bar_color = "#16A34A" if pct >= 70 else "#F39C12" if pct >= 45 else "#DC2626"
            st.markdown(f"**{ss.name}** ({ss.weight*100:.0f}% weight) — {pct}/100")
            st.markdown(
                f"<div style='background:#F1F5F9;border-radius:4px;height:10px;border:1px solid #E2E8F0'>"
                f"<div style='background:{bar_color};width:{pct}%;height:10px;border-radius:4px'></div></div>",
                unsafe_allow_html=True,
            )
            st.markdown(f"<small style='color:#6B7280'>{ss.detail}</small>", unsafe_allow_html=True)
            st.markdown("")

    with col_ai:
        st.markdown('<div class="section-header">🤖 AI-Assisted Decision</div>', unsafe_allow_html=True)
        rec_color = {"APPROVE": "#2ECC71", "CONDITIONAL": "#F39C12", "REJECT": "#DC2626"}.get(
            ai.recommendation, "#888"
        )
        st.markdown(
            f"<div style='font-size:2rem;font-weight:800;color:{rec_color}'>"
            f"{ai.recommendation}</div>",
            unsafe_allow_html=True,
        )
        st.markdown(
            f"Confidence: **{ai.confidence_level}** ({ai.confidence_score:.0%})  "
            f"| Model: `{ai.model_used}`"
        )
        st.progress(ai.confidence_score)

        st.markdown("**AI Assessment**")
        # Render narrative — already bullet-formatted by the prompt
        # Split on bullet character to style each point individually
        narrative_lines = [
            ln.strip().lstrip("•").strip()
            for ln in ai.risk_narrative.split("\n")
            if ln.strip().lstrip("•").strip()
        ]
        for line in narrative_lines:
            st.markdown(
                f'<div style="background:#F0F7FF;border-left:3px solid #2274E0;'
                f'padding:7px 12px;border-radius:5px;margin:4px 0;font-size:0.9rem">'
                f'• {line}</div>',
                unsafe_allow_html=True,
            )

        if ai.key_risks:
            st.markdown("**Key Risks**")
            for risk in ai.key_risks:
                st.markdown(
                    f'<div class="red-flag">⚠️ {risk}</div>',
                    unsafe_allow_html=True,
                )

        if ai.positive_factors:
            st.markdown("**Positive Factors**")
            for pos in ai.positive_factors:
                st.markdown(
                    f'<div class="pos-signal">✅ {pos}</div>',
                    unsafe_allow_html=True,
                )

        if ai.hidden_risks:
            with st.expander("🔍 Hidden / Structural Risks"):
                for hr in ai.hidden_risks:
                    st.markdown(
                        f'<div style="background:#F0F7FF;border-left:3px solid #555;'
                        f'padding:6px 10px;border-radius:4px;margin:3px 0;font-size:0.85rem">'
                        f'⚠️ {hr}</div>',
                        unsafe_allow_html=True,
                    )

    # Credit decision strip
    st.markdown('<div class="section-header" style="margin-top:8px">💳 Credit Decision Summary</div>', unsafe_allow_html=True)
    dc1, dc2, dc3, dc4 = st.columns(4)
    approve_icon = "✅" if decision.approved else "❌"
    dc1.metric(f"{approve_icon} Decision",    "APPROVED" if decision.approved else "DECLINED")
    dc2.metric("Credit Limit",                fmt_currency(decision.credit_limit))
    dc3.metric("Max Safe Exposure",           fmt_currency(decision.max_safe_exposure))
    dc4.metric("Payment Terms",               f"Net-{decision.payment_terms_days}")

    st.markdown("**Limit Calculation Breakdown**")
    st.markdown(decision.rationale)

    if decision.conditions:
        st.markdown("**Conditions / Covenants**")
        for cond in decision.conditions:
            st.markdown(
                f'<div class="cond-item">📋 {cond}</div>',
                unsafe_allow_html=True,
            )

    if ai.suggested_conditions:
        st.markdown("**AI-Suggested Conditions**")
        for cond in ai.suggested_conditions:
            st.markdown(
                f'<div class="cond-item">🤖 {cond}</div>',
                unsafe_allow_html=True,
            )


def render_alerts(r: dict):
    scoring = r["scoring"]
    ai      = r["ai"]

    all_flags = list(scoring.red_flags) + [f"AI: {rk}" for rk in ai.key_risks[:3]]

    st.markdown('<div class="section-header">🚨 Alerts &amp; Red Flags</div>', unsafe_allow_html=True)

    if not all_flags:
        st.success("✅ No critical alerts — risk profile is clean.")
    else:
        st.markdown('<div class="panel-title">🔴 Active Risk Flags</div>', unsafe_allow_html=True)
        flags_html = "".join(
            f'<div class="red-flag" style="font-size:1rem">🔴 {flag}</div>'
            for flag in all_flags
        )
        st.markdown(flags_html, unsafe_allow_html=True)

    if scoring.positive_signals:
        st.markdown('<hr class="sec-divider">', unsafe_allow_html=True)
        st.markdown('<div class="panel-title">✅ Positive Signals</div>', unsafe_allow_html=True)
        sigs_html = "".join(
            f'<div class="pos-signal">✅ {sig}</div>'
            for sig in scoring.positive_signals[:5]
        )
        st.markdown(sigs_html, unsafe_allow_html=True)


def render_explainability(r: dict):
    scoring  = r["scoring"]
    decision = r["decision"]

    st.markdown('<div class="section-header">🔍 Score Explainability</div>', unsafe_allow_html=True)

    st.markdown('<div class="section-header" style="background:#F0F7FF;margin-top:4px">🧮 How the Composite Score is Computed</div>', unsafe_allow_html=True)

    _sub_inputs = {
        "Financial Health":      "8 ratios incl. DSO, normalised to 0–100",
        "Payment Behavior":      "Historical invoice ageing buckets",
        "Financial Trend":       "YoY revenue growth, margin trend, debt trajectory",
        "External Risk":         "News sentiment + risk flag count",
        "Operational Stability": "Revenue, cash buffer, external flags",
    }
    has_pay = any(ss.name == "Payment Behavior" and ss.weight > 0 for ss in scoring.sub_scores)
    if not has_pay:
        st.info(
            "**Payment history not provided.** The 25% Payment Behavior weight has been "
            "redistributed proportionally among the other four sub-models for this evaluation.",
            icon="ℹ️",
        )

    weight_rows = []
    for ss in scoring.sub_scores:
        weight_rows.append({
            "Sub-Model": ss.name,
            "Weight":    f"{ss.weight * 100:.0f}%",
            "Score":     f"{ss.score:.0f}/100",
            "Contribution": f"{ss.weighted_contribution:.1f} pts",
            "Inputs":    _sub_inputs.get(ss.name, ""),
        })
    st.dataframe(pd.DataFrame(weight_rows), use_container_width=True, hide_index=True)

    st.markdown("```\nComposite = Σ (sub_score × weight)\n```")
    st.markdown('<hr class="sec-divider">', unsafe_allow_html=True)

    # Ratio benchmark table
    st.markdown('<div class="section-header" style="background:#F0F7FF;margin-top:4px">📐 Financial Ratio Benchmarks</div>', unsafe_allow_html=True)
    bm_rows = []
    for rr in scoring.ratios:
        from config import RATIO_BENCHMARKS
        bm = RATIO_BENCHMARKS[rr.name]
        bm_rows.append({
            "Ratio":            rr.name.replace("_", " ").title(),
            "Your Value":       rr.display_value,
            "Score (0-100)":    f"{rr.normalised_score:.0f}",
            "Low-Risk Min":     str(bm.low_risk_min),
            "Acceptable Min":   str(bm.acceptable_min),
            "Direction":        "↑ Higher better" if bm.higher_is_better else "↓ Lower better",
            "Signal":           rr.signal.upper(),
        })
    st.dataframe(pd.DataFrame(bm_rows), use_container_width=True, hide_index=True)
    st.markdown('<hr class="sec-divider">', unsafe_allow_html=True)

    # Credit limit formula
    st.markdown('<div class="section-header" style="background:#F0F7FF;margin-top:4px">💳 Credit Limit Formula</div>', unsafe_allow_html=True)
    dso_str = f"{decision.dso_days:.0f} days" if decision.dso_days is not None else "N/A (no AR data)"
    st.markdown(f"""
    ```
    Travel Budget       = {fmt_currency(decision.travel_budget)}/yr
                          ({decision.travel_budget_source})
    Monthly Travel      = {fmt_currency(decision.monthly_travel)}
    Base Limit          = {fmt_currency(decision.monthly_travel)} × 3 months
                        = {fmt_currency(decision.base_limit)}

    Cash Flow Mult      = {decision.cash_flow_multiplier:.2f}x   (OCF/Debt coverage)
    Score Mult          = {decision.score_multiplier:.2f}x        (Credit score {decision.composite_score:.0f}/100)
    Sentiment Mult      = {decision.sentiment_multiplier:.2f}x   (External news)
    DSO Mult            = {decision.dso_multiplier:.2f}x          (Receivables: {dso_str})

    Final Multiplier    = {decision.cash_flow_multiplier:.2f} × {decision.score_multiplier:.2f}
                          × {decision.sentiment_multiplier:.2f} × {decision.dso_multiplier:.2f}
                        = {decision.final_multiplier:.3f}x

    Recommended Limit   = {fmt_currency(decision.base_limit)} × {decision.final_multiplier:.3f}
                        = {fmt_currency(decision.credit_limit)}

    Max Safe Exposure   = {fmt_currency(decision.credit_limit)} × 1.25
                        = {fmt_currency(decision.max_safe_exposure)}
    ```
    """)
    st.markdown('<hr class="sec-divider">', unsafe_allow_html=True)

    # Waterfall chart of sub-score contributions
    st.markdown('<div class="section-header" style="background:#F0F7FF;margin-top:4px">📉 Score Contribution Waterfall</div>', unsafe_allow_html=True)
    labels = list(scoring.score_breakdown.keys())
    values = list(scoring.score_breakdown.values())

    fig_wf = go.Figure(go.Waterfall(
        name="",
        orientation="v",
        measure=["relative"] * (len(labels) - 1) + ["total"],
        x=labels,
        y=values,
        connector={"line": {"color": "rgb(63, 63, 63)"}},
        increasing={"marker": {"color": "#16A34A"}},
        decreasing={"marker": {"color": "#DC2626"}},
        totals={"marker": {"color": "#2274E0"}},
    ))
    fig_wf.update_layout(
        height=350, paper_bgcolor="#FFFFFF", font_color="#0F172A", showlegend=False,
        yaxis_title="Score Points",
        margin=dict(l=10, r=10, t=10, b=10),
    )
    st.plotly_chart(fig_wf, use_container_width=True)


# ---------------------------------------------------------------------------
# History full-page view
# ---------------------------------------------------------------------------

def _db_row_to_result(db: dict) -> dict:
    """
    Reconstruct the live result dict (with proper dataclass objects) from a
    database row returned by get_evaluation().  This allows historical records
    to be rendered by all the same panel functions.
    """
    from scoring import ScoringResult, SubScoreResult, RatioResult
    from ai_engine import AIDecision
    from credit_engine import CreditDecision

    # ── Financials & payment ───────────────────────────────────────────────
    financials      = db.get("financials") or {}
    payment_records = (db.get("payment") or {}).get("records", [])
    external        = db.get("external") or {}

    # ── Rebuild ScoringResult ──────────────────────────────────────────────
    sc = db.get("scoring") or {}

    ratio_objects = [
        RatioResult(
            name                = r.get("name", ""),
            value               = r.get("value"),
            normalised_score    = r.get("normalised_score", 50.0),
            benchmark_low_risk  = r.get("benchmark_low_risk", 0),
            benchmark_acceptable= r.get("benchmark_acceptable", 0),
            higher_is_better    = r.get("higher_is_better", True),
            display_value       = r.get("display_value", "N/A"),
            signal              = r.get("signal", "ok"),
            explanation         = r.get("explanation", ""),
        )
        for r in (sc.get("ratios") or [])
    ]

    subscore_objects = [
        SubScoreResult(
            name                  = s.get("name", ""),
            score                 = s.get("score", 50.0),
            weight                = s.get("weight", 0.0),
            weighted_contribution = s.get("weighted_contribution", 0.0),
            detail                = s.get("detail", ""),
        )
        for s in (sc.get("sub_scores") or [])
    ]

    scoring_obj = ScoringResult(
        composite_score  = sc.get("composite_score", 50.0),
        risk_category    = sc.get("risk_category", "Average"),
        risk_color       = sc.get("risk_color", "#F39C12"),
        risk_emoji       = sc.get("risk_emoji", "🟡"),
        ratios           = ratio_objects,
        sub_scores       = subscore_objects,
        score_breakdown  = sc.get("score_breakdown", {}),
        red_flags        = sc.get("red_flags", []),
        positive_signals = sc.get("positive_signals", []),
    )

    # ── Rebuild AIDecision ─────────────────────────────────────────────────
    narrative_raw = db.get("ai_narrative") or "No AI narrative saved."
    # Parse from stored JSON if available (newer saves store full AI JSON in narrative)
    ai_json = {}
    try:
        import json as _json, re as _re
        m = _re.search(r"\{.*\}", narrative_raw, _re.DOTALL)
        if m:
            ai_json = _json.loads(m.group())
    except Exception:
        pass

    ai_obj = AIDecision(
        recommendation       = ai_json.get("recommendation", sc.get("risk_category", "CONDITIONAL")),
        confidence_level     = ai_json.get("confidence_level", "Medium"),
        confidence_score     = float(db.get("ai_confidence") or ai_json.get("confidence_score", 0.6)),
        risk_narrative       = ai_json.get("risk_narrative", narrative_raw),
        key_risks            = ai_json.get("key_risks", sc.get("red_flags", [])[:4]),
        positive_factors     = ai_json.get("positive_factors", sc.get("positive_signals", [])[:4]),
        hidden_risks         = ai_json.get("hidden_risks", []),
        suggested_conditions = ai_json.get("suggested_conditions", []),
        model_used           = ai_json.get("model_used", "stored"),
    )

    # ── Rebuild CreditDecision ─────────────────────────────────────────────
    dec = db.get("decision") or {}

    decision_obj = CreditDecision(
        approved             = dec.get("approved", False),
        credit_limit         = dec.get("credit_limit", 0.0),
        max_safe_exposure    = dec.get("max_safe_exposure", 0.0),
        payment_terms_days   = dec.get("payment_terms_days", dec.get("payment_terms", 30)),
        risk_category        = dec.get("risk_category", "Average"),
        composite_score      = dec.get("composite_score", scoring_obj.composite_score),
        travel_budget        = dec.get("travel_budget", 0.0),
        monthly_travel       = dec.get("monthly_travel", 0.0),
        travel_budget_source = dec.get("travel_budget_source", "stored"),
        travel_budget_pct    = dec.get("travel_budget_pct", 0.025),
        dso_days             = dec.get("dso_days"),
        dso_multiplier       = dec.get("dso_multiplier", 1.0),
        base_limit           = dec.get("base_limit", 0.0),
        monthly_revenue      = dec.get("monthly_revenue", 0.0),
        cash_flow_multiplier = dec.get("cash_flow_multiplier", 1.0),
        score_multiplier     = dec.get("score_multiplier", 1.0),
        sentiment_multiplier = dec.get("sentiment_multiplier", 1.0),
        final_multiplier     = dec.get("final_multiplier", 1.0),
        rationale            = dec.get("rationale", ""),
        conditions           = dec.get("conditions", []),
    )

    return {
        "eval_id":          db.get("id"),
        "company_name":     db.get("company_name", ""),
        "financials":       financials,
        "payment_records":  payment_records,
        "external":         external,
        "scoring":          scoring_obj,
        "ai":               ai_obj,
        "decision":         decision_obj,
        "parse_confidence": db.get("parse_confidence", 1.0),
        "parse_warnings":   [],
    }


def render_history():
    from config import RISK_COLORS, RISK_EMOJI

    st.markdown("## 🗂 Evaluation History")

    query       = st.session_state.get("hist_search_query", "")
    risk_filter = st.session_state.get("hist_risk_filter", "All")
    rows = search_evaluations(query=query, risk_filter=risk_filter, limit=200)

    # ── Summary metrics ────────────────────────────────────────────────────
    all_rows = list_evaluations(limit=1000)
    total        = len(all_rows)
    approved     = sum(1 for r in all_rows if r.get("risk_category") in ("Low Risk", "Average"))
    avg_score    = round(sum(r.get("composite_score", 0) for r in all_rows) / total, 1) if total else 0
    unique_cos   = len({r["company_name"] for r in all_rows})

    m1, m2, m3, m4 = st.columns(4)
    m1.metric("Total Evaluations", total)
    m2.metric("Unique Companies",  unique_cos)
    m3.metric("Avg Composite Score", f"{avg_score}/100")
    m4.metric("Approved / Conditional", approved)

    st.divider()

    if not rows:
        st.info("No evaluations match your search. Try a different name or clear the filter.")
        return

    # ── Build display dataframe ────────────────────────────────────────────
    _RISK_EMOJI = {"Low Risk": "🟢", "Average": "🟡", "Risky": "🔴"}

    display_rows = []
    for r in rows:
        score = r.get("composite_score") or 0
        risk  = r.get("risk_category") or "—"
        limit_val = r.get("credit_limit") or 0
        display_rows.append({
            "ID":             r["id"],
            "Company":        r["company_name"],
            "Date":           (r.get("created_at") or "")[:16].replace("T", "  "),
            "Score":          f"{score:.0f}",
            "Risk":           f"{_RISK_EMOJI.get(risk, '⚪')} {risk}",
            "Credit Limit":   fmt_currency(limit_val),
            "Terms":          f"Net-{r.get('payment_terms', '—')}",
            "Source":         r.get("source_file") or "Manual",
        })

    df_display = pd.DataFrame(display_rows)

    # ── Colour-coded table via st.dataframe ────────────────────────────────
    st.markdown(f"**{len(rows)} evaluation(s) found**")

    # Use Streamlit's built-in column config for a polished look
    col_cfg = {
        "ID":           st.column_config.NumberColumn("ID",    width="small"),
        "Company":      st.column_config.TextColumn("Company", width="medium"),
        "Date":         st.column_config.TextColumn("Date",    width="medium"),
        "Score":        st.column_config.TextColumn("Score",   width="small"),
        "Risk":         st.column_config.TextColumn("Risk Category", width="medium"),
        "Credit Limit": st.column_config.TextColumn("Credit Limit",  width="medium"),
        "Terms":        st.column_config.TextColumn("Terms",  width="small"),
        "Source":       st.column_config.TextColumn("Source", width="medium"),
    }

    selected = st.dataframe(
        df_display,
        use_container_width=True,
        hide_index=True,
        column_config=col_cfg,
        on_select="rerun",
        selection_mode="single-row",
        key="history_table",
    )

    # ── Load selected row ──────────────────────────────────────────────────
    sel_rows = (selected.get("selection", {}) or {}).get("rows", [])
    if sel_rows:
        chosen_idx = sel_rows[0]
        chosen_id  = display_rows[chosen_idx]["ID"]
        chosen_co  = display_rows[chosen_idx]["Company"]

        st.info(f"**Selected:** {chosen_co} (ID {chosen_id})")
        if st.button("📂 Load This Evaluation", type="primary", use_container_width=False):
            full = get_evaluation(chosen_id)
            if full:
                st.session_state["result"] = _db_row_to_result(full)
                # Switch back to result view; reset any stale email dialog state
                st.session_state["sidebar_mode_override"] = True
                st.session_state["open_email_dlg"] = False
                st.rerun()
            else:
                st.error("Could not load evaluation — record may have been deleted.")
    else:
        st.caption("Click a row in the table above to select it, then press Load.")

    # ── Quick delete (optional, power-user) ───────────────────────────────
    with st.expander("⚠️ Danger Zone — Delete Records"):
        del_id = st.number_input("Enter evaluation ID to delete", min_value=1, step=1, key="del_eval_id")
        if st.button("Delete", type="secondary"):
            try:
                del_id_int = int(del_id)
                from database import _conn as _db_conn
                with _db_conn() as con:
                    cur = con.execute(
                        "DELETE FROM evaluations WHERE id = ?", (del_id_int,)
                    )
                    deleted = cur.rowcount

                if deleted == 0:
                    st.warning(
                        f"No evaluation found with ID **{del_id_int}**. "
                        "Please check the ID in the table above and try again."
                    )
                else:
                    st.success(f"Evaluation ID {del_id_int} deleted successfully.")
                    st.rerun()

            except ValueError:
                st.error("Invalid ID — please enter a whole number.")
            except Exception as e:
                st.error(f"Delete failed: {e}")


# ---------------------------------------------------------------------------
# Main layout
# ---------------------------------------------------------------------------

def main():
    render_sidebar()

    if st.session_state["running"]:
        st.info("Running evaluation…")
        return

    # Editable extraction preview (two-step image flow)
    if st.session_state.get("awaiting_review"):
        _render_extraction_editor()
        return

    # If History mode is active, show history panel — UNLESS a result has been
    # loaded from history (sidebar_mode_override) OR there is already a live
    # result in session state to display.
    sidebar_mode = st.session_state.get("sidebar_mode", "Image / File Upload")
    result = st.session_state.get("result")

    if sidebar_mode == "History" and result is None and not st.session_state.get("sidebar_mode_override"):
        render_history()
        return

    # Clear the override flag only when the user explicitly switches away from
    # History mode (i.e., the sidebar selector is no longer "History").
    if sidebar_mode != "History":
        st.session_state.pop("sidebar_mode_override", None)

    if result is None:
        # Welcome screen
        import os as _os
        if not _os.environ.get("ANTHROPIC_API_KEY"):
            st.warning(
                "### ⚡ Claude Vision is not configured\n\n"
                "This dashboard uses **Claude Vision** (the same AI as claude.ai) to read "
                "financial table images accurately. Without an API key it falls back to basic OCR "
                "which often misses values.\n\n"
                "**To enable Claude Vision:**\n"
                "1. Go to **console.anthropic.com** → API Keys → Create Key\n"
                "2. Paste the key in the **sidebar** (Image / File Upload → API Key field)\n"
                "3. Re-upload your image — it will be read perfectly"
            )
        st.markdown(f"# {APP_ICON} {APP_TITLE}")
        st.markdown("### B2B Travel Credit Risk Platform")
        st.markdown("""
        Upload an **image** (photo/screenshot) of a financial statement, enter figures
        **manually**, or reload a past evaluation from **History**:

        - **Image extraction** — Claude Vision reads P&L, balance sheet, or annual report images
        - **Dual scoring** — rule-based weighted model + AI narrative
        - **Dynamic credit limit** — travel-anchored, risk-adjusted exposure
        - **Live news signals** — real-time sentiment via Google News RSS
        - **Full explainability** — every formula, benchmark and weight shown

        **Get started in the sidebar →**
        """)

        # Quick-start with sample data
        st.markdown("---")
        st.subheader("Quick Demo with Sample Data")
        col_demo1, col_demo2, col_demo3 = st.columns(3)

        if col_demo1.button("Run GOOD Company Demo", use_container_width=True):
            _load_demo("good")
        if col_demo2.button("Run AVERAGE Company Demo", use_container_width=True):
            _load_demo("average")
        if col_demo3.button("Run RISKY Company Demo", use_container_width=True):
            _load_demo("risky")
        return

    # --- Build tabs for the result ---
    tabs = st.tabs([
        "📊 Snapshot",
        "📈 Financials",
        "⏱ Payments",
        "🌐 External Risk",
        "🧠 Decisions",
        "🚨 Alerts",
        "🔍 Explainability",
    ])

    with tabs[0]: render_snapshot(result)
    with tabs[1]: render_financial_health(result)
    with tabs[2]: render_payment_behavior(result)
    with tabs[3]: render_external_risk(result)
    with tabs[4]: render_decisions(result)
    with tabs[5]: render_alerts(result)
    with tabs[6]: render_explainability(result)


# ---------------------------------------------------------------------------
# Demo profiles
# ---------------------------------------------------------------------------

_DEMO_PROFILES = {
    # ₹ 100 Cr revenue — strong corporate client
    "good": {
        "company_name": "Apex Logistics International Pvt. Ltd.",
        "revenue":             100_00_00_000,   # ₹ 100 Cr
        "ebitda":               20_00_00_000,   # ₹  20 Cr
        "net_profit":           11_50_00_000,   # ₹  11.5 Cr
        "current_assets":       37_50_00_000,   # ₹  37.5 Cr
        "current_liabilities":  15_00_00_000,   # ₹  15 Cr
        "total_debt":           25_00_00_000,   # ₹  25 Cr
        "equity":               50_00_00_000,   # ₹  50 Cr
        "cash_and_equivalents":  7_50_00_000,   # ₹   7.5 Cr
        "operating_cash_flow":  15_00_00_000,   # ₹  15 Cr
        "interest_expense":      2_00_00_000,   # ₹   2 Cr
    },
    # ₹ 60 Cr revenue — mid-tier, mixed signals
    "average": {
        "company_name": "Meridian Travel Solutions Ltd.",
        "revenue":             60_00_00_000,    # ₹ 60 Cr
        "ebitda":               6_00_00_000,    # ₹  6 Cr
        "net_profit":           2_25_00_000,    # ₹  2.25 Cr
        "current_assets":      17_50_00_000,    # ₹ 17.5 Cr
        "current_liabilities": 14_50_00_000,    # ₹ 14.5 Cr
        "total_debt":          35_00_00_000,    # ₹ 35 Cr
        "equity":              22_50_00_000,    # ₹ 22.5 Cr
        "cash_and_equivalents": 2_00_00_000,    # ₹  2 Cr
        "operating_cash_flow":  4_50_00_000,    # ₹  4.5 Cr
        "interest_expense":     2_50_00_000,    # ₹  2.5 Cr
    },
    # ₹ 25 Cr revenue — distressed, rejection scenario
    "risky": {
        "company_name": "Zenith Corporate Flights Pvt. Ltd.",
        "revenue":             25_00_00_000,    # ₹ 25 Cr
        "ebitda":                -65_00_000,    # ₹ -0.65 Cr
        "net_profit":          -2_80_00_000,    # ₹ -2.8 Cr
        "current_assets":       6_50_00_000,    # ₹  6.5 Cr
        "current_liabilities":  9_75_00_000,    # ₹  9.75 Cr
        "total_debt":          31_00_00_000,    # ₹ 31 Cr
        "equity":               3_25_00_000,    # ₹  3.25 Cr
        "cash_and_equivalents":   65_00_000,    # ₹  0.65 Cr
        "operating_cash_flow":   -97_50_000,    # ₹ -0.975 Cr
        "interest_expense":     2_25_00_000,    # ₹  2.25 Cr
    },
}


def _load_demo(profile: str):
    fin = _DEMO_PROFILES[profile].copy()
    company_name = fin["company_name"]
    _run_evaluation(
        fin_bytes=None, filename="", company_name=company_name,
        pay_bytes=None, use_mock_news=False,
        manual_financials=fin,
    )


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    main()





